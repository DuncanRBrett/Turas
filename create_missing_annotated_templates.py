#!/usr/bin/env python3
"""
Create annotated TURAS module templates
Generates annotated Pricing, KeyDriver, and Conjoint configuration templates
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Shared styling
BLUE_HEADER = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
YELLOW_DOC = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
GRAY_EXAMPLE = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
BOLD_WHITE_FONT = Font(bold=True, color='FFFFFF', size=11)
BOLD_BLACK_FONT = Font(bold=True, size=11)


def create_pricing_annotated(output_file):
    """Create Pricing_Config_Template_Annotated.xlsx"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ===== Instructions Sheet =====
    ws_inst = wb.create_sheet("Instructions", 0)
    ws_inst['A1'] = 'TURAS Pricing Analysis - Configuration Template'
    ws_inst['A1'].font = Font(bold=True, size=14, color='4472C4')

    instructions = [
        '',
        'This template configures pricing research analysis using Van Westendorp and/or Gabor-Granger methods.',
        '',
        'SHEETS IN THIS TEMPLATE:',
        '  • Settings - Core analysis parameters',
        '  • VanWestendorp - PSM question mapping (if using Van Westendorp)',
        '  • GaborGranger - Price points and purchase intent columns (if using Gabor-Granger)',
        '',
        'SUPPORTED METHODS:',
        '  • van_westendorp - Price Sensitivity Meter analysis',
        '  • gabor_granger - Revenue maximization analysis',
        '  • both - Run both methods on same dataset',
        '',
        'QUICK START:',
        '  1. Update Settings sheet with your data file path and method',
        '  2. If using Van Westendorp, map your PSM questions in VanWestendorp sheet',
        '  3. If using Gabor-Granger, define price points in GaborGranger sheet',
        '  4. Save and run: turas_load("pricing") then run analysis',
        '',
        'DOCUMENTATION:',
        '  • Yellow rows = parameter documentation (Required?, Valid Values, Description)',
        '  • Gray rows = example values',
        '  • Delete documentation rows before production use (optional)',
        '',
        'For detailed help, see: modules/pricing/documentation/'
    ]

    for i, text in enumerate(instructions, start=2):
        ws_inst[f'A{i}'] = text

    ws_inst.column_dimensions['A'].width = 90

    # ===== Settings Sheet =====
    ws_settings = wb.create_sheet("Settings")

    # Headers
    ws_settings['A1'] = 'Setting'
    ws_settings['B1'] = 'Value'
    for cell in ['A1', 'B1']:
        ws_settings[cell].font = BOLD_WHITE_FONT
        ws_settings[cell].fill = BLUE_HEADER

    row = 2

    # analysis_method
    ws_settings[f'A{row}'] = 'analysis_method'
    ws_settings[f'A{row}'].font = BOLD_BLACK_FONT
    row += 1
    doc_cells = [f'A{row}', f'B{row}']
    ws_settings[f'A{row}'] = 'Required?'
    ws_settings[f'B{row}'] = 'Valid Values'
    for cell in doc_cells:
        ws_settings[cell].fill = YELLOW_DOC
        ws_settings[cell].font = Font(italic=True, size=9)
    row += 1
    doc_cells = [f'A{row}', f'B{row}']
    ws_settings[f'A{row}'] = 'Required'
    ws_settings[f'B{row}'] = 'van_westendorp | gabor_granger | both'
    for cell in doc_cells:
        ws_settings[cell].fill = YELLOW_DOC
        ws_settings[cell].font = Font(italic=True, size=9)
    row += 1
    ws_settings[f'A{row}'] = 'Description'
    ws_settings[f'A{row}'].fill = YELLOW_DOC
    ws_settings[f'A{row}'].font = Font(italic=True, size=9)
    ws_settings.merge_cells(f'B{row}:B{row}')
    ws_settings[f'B{row}'] = 'Analysis method: Van Westendorp PSM, Gabor-Granger revenue optimization, or both'
    ws_settings[f'B{row}'].fill = YELLOW_DOC
    ws_settings[f'B{row}'].alignment = Alignment(wrap_text=True)
    ws_settings[f'B{row}'].font = Font(italic=True, size=9)
    row += 1
    ws_settings[f'A{row}'] = 'analysis_method'
    ws_settings[f'B{row}'] = 'van_westendorp'
    ws_settings[f'A{row}'].fill = GRAY_EXAMPLE
    ws_settings[f'B{row}'].fill = GRAY_EXAMPLE
    row += 1
    row += 1

    # data_file
    ws_settings[f'A{row}'] = 'data_file'
    ws_settings[f'A{row}'].font = BOLD_BLACK_FONT
    row += 1
    doc_cells = [f'A{row}', f'B{row}']
    ws_settings[f'A{row}'] = 'Required'
    ws_settings[f'B{row}'] = 'CSV, XLSX, SAV, RDS file path'
    for cell in doc_cells:
        ws_settings[cell].fill = YELLOW_DOC
        ws_settings[cell].font = Font(italic=True, size=9)
    row += 1
    ws_settings[f'A{row}'] = 'Description'
    ws_settings.merge_cells(f'B{row}:B{row}')
    ws_settings[f'B{row}'] = 'Path to data file (relative to config file or absolute)'
    ws_settings[f'A{row}'].fill = YELLOW_DOC
    ws_settings[f'B{row}'].fill = YELLOW_DOC
    ws_settings[f'A{row}'].font = Font(italic=True, size=9)
    ws_settings[f'B{row}'].font = Font(italic=True, size=9)
    row += 1
    ws_settings[f'A{row}'] = 'data_file'
    ws_settings[f'B{row}'] = 'pricing_data.csv'
    ws_settings[f'A{row}'].fill = GRAY_EXAMPLE
    ws_settings[f'B{row}'].fill = GRAY_EXAMPLE
    row += 1
    row += 1

    # output_file
    ws_settings[f'A{row}'] = 'output_file'
    ws_settings[f'A{row}'].font = BOLD_BLACK_FONT
    row += 1
    doc_cells = [f'A{row}', f'B{row}']
    ws_settings[f'A{row}'] = 'Required'
    ws_settings[f'B{row}'] = 'XLSX file path'
    for cell in doc_cells:
        ws_settings[cell].fill = YELLOW_DOC
        ws_settings[cell].font = Font(italic=True, size=9)
    row += 1
    ws_settings[f'A{row}'] = 'Description'
    ws_settings.merge_cells(f'B{row}:B{row}')
    ws_settings[f'B{row}'] = 'Output Excel file path (relative to config file or absolute)'
    ws_settings[f'A{row}'].fill = YELLOW_DOC
    ws_settings[f'B{row}'].fill = YELLOW_DOC
    ws_settings[f'A{row}'].font = Font(italic=True, size=9)
    ws_settings[f'B{row}'].font = Font(italic=True, size=9)
    row += 1
    ws_settings[f'A{row}'] = 'output_file'
    ws_settings[f'B{row}'] = 'pricing_results.xlsx'
    ws_settings[f'A{row}'].fill = GRAY_EXAMPLE
    ws_settings[f'B{row}'].fill = GRAY_EXAMPLE
    row += 1
    row += 1

    # Optional parameters
    optional_params = [
        ['weight_var', 'Optional', 'Column name', 'Weight variable name for weighted analysis'],
        ['dk_codes', 'Optional', 'Comma-separated numbers', "Don't know codes to exclude (e.g., 98,99)"],
        ['vw_monotonicity_behavior', 'Optional', 'flag_only | drop | fix', 'How to handle non-monotonic VW responses (default: flag_only)'],
        ['gg_monotonicity_behavior', 'Optional', 'smooth | flag_only | none', 'How to handle non-monotonic GG curves (default: smooth)'],
        ['segment_vars', 'Optional', 'Comma-separated column names', 'Variables to segment by (e.g., age_group,region)'],
        ['unit_cost', 'Optional', 'Positive number', 'Unit cost for profit optimization in Gabor-Granger']
    ]

    for param_info in optional_params:
        param, required, valid, desc = param_info
        ws_settings[f'A{row}'] = param
        ws_settings[f'A{row}'].font = BOLD_BLACK_FONT
        row += 1
        ws_settings[f'A{row}'] = required
        ws_settings[f'B{row}'] = valid
        ws_settings[f'A{row}'].fill = YELLOW_DOC
        ws_settings[f'B{row}'].fill = YELLOW_DOC
        ws_settings[f'A{row}'].font = Font(italic=True, size=9)
        ws_settings[f'B{row}'].font = Font(italic=True, size=9)
        row += 1
        ws_settings[f'A{row}'] = 'Description'
        ws_settings.merge_cells(f'B{row}:B{row}')
        ws_settings[f'B{row}'] = desc
        ws_settings[f'A{row}'].fill = YELLOW_DOC
        ws_settings[f'B{row}'].fill = YELLOW_DOC
        ws_settings[f'A{row}'].font = Font(italic=True, size=9)
        ws_settings[f'B{row}'].font = Font(italic=True, size=9)
        ws_settings[f'B{row}'].alignment = Alignment(wrap_text=True)
        row += 1
        ws_settings[f'A{row}'] = param
        ws_settings[f'B{row}'] = ''
        ws_settings[f'A{row}'].fill = GRAY_EXAMPLE
        ws_settings[f'B{row}'].fill = GRAY_EXAMPLE
        row += 1
        row += 1

    ws_settings.column_dimensions['A'].width = 30
    ws_settings.column_dimensions['B'].width = 50

    # ===== VanWestendorp Sheet =====
    ws_vw = wb.create_sheet("VanWestendorp")

    ws_vw['A1'] = 'Question'
    ws_vw['B1'] = 'ColumnName'
    for cell in ['A1', 'B1']:
        ws_vw[cell].font = BOLD_WHITE_FONT
        ws_vw[cell].fill = BLUE_HEADER

    row = 2
    vw_questions = [
        ['too_cheap', 'Price point where product seems too cheap (quality concern)', 'vw_too_cheap'],
        ['cheap', 'Price point where product is a bargain', 'vw_cheap'],
        ['expensive', 'Price point where product starts getting expensive', 'vw_expensive'],
        ['too_expensive', 'Price point where product is too expensive to consider', 'vw_too_expensive']
    ]

    for q_type, desc, example_col in vw_questions:
        ws_vw[f'A{row}'] = q_type
        ws_vw[f'A{row}'].font = BOLD_BLACK_FONT
        row += 1
        ws_vw[f'A{row}'] = 'Description'
        ws_vw.merge_cells(f'B{row}:B{row}')
        ws_vw[f'B{row}'] = desc
        ws_vw[f'A{row}'].fill = YELLOW_DOC
        ws_vw[f'B{row}'].fill = YELLOW_DOC
        ws_vw[f'A{row}'].font = Font(italic=True, size=9)
        ws_vw[f'B{row}'].font = Font(italic=True, size=9)
        ws_vw[f'B{row}'].alignment = Alignment(wrap_text=True)
        row += 1
        ws_vw[f'A{row}'] = q_type
        ws_vw[f'B{row}'] = example_col
        ws_vw[f'A{row}'].fill = GRAY_EXAMPLE
        ws_vw[f'B{row}'].fill = GRAY_EXAMPLE
        row += 1
        row += 1

    ws_vw.column_dimensions['A'].width = 20
    ws_vw.column_dimensions['B'].width = 50

    # ===== GaborGranger Sheet =====
    ws_gg = wb.create_sheet("GaborGranger")

    ws_gg['A1'] = 'PricePoint'
    ws_gg['B1'] = 'PurchaseIntentColumn'
    for cell in ['A1', 'B1']:
        ws_gg[cell].font = BOLD_WHITE_FONT
        ws_gg[cell].fill = BLUE_HEADER

    # Documentation rows
    ws_gg['A2'] = 'Required?'
    ws_gg['B2'] = 'Valid Values'
    ws_gg['A2'].fill = YELLOW_DOC
    ws_gg['B2'].fill = YELLOW_DOC
    ws_gg['A2'].font = Font(italic=True, size=9)
    ws_gg['B2'].font = Font(italic=True, size=9)

    ws_gg['A3'] = 'Required (3+ points)'
    ws_gg['B3'] = 'Column name with 0/1 purchase intent'
    ws_gg['A3'].fill = YELLOW_DOC
    ws_gg['B3'].fill = YELLOW_DOC
    ws_gg['A3'].font = Font(italic=True, size=9)
    ws_gg['B3'].font = Font(italic=True, size=9)

    ws_gg['A4'] = 'Description'
    ws_gg.merge_cells('B4:B4')
    ws_gg['B4'] = 'Define tested price points and corresponding purchase intent columns. Need minimum 3 price points.'
    ws_gg['A4'].fill = YELLOW_DOC
    ws_gg['B4'].fill = YELLOW_DOC
    ws_gg['A4'].font = Font(italic=True, size=9)
    ws_gg['B4'].font = Font(italic=True, size=9)
    ws_gg['B4'].alignment = Alignment(wrap_text=True)

    # Example data
    gg_examples = [
        [49, 'pi_49'],
        [69, 'pi_69'],
        [89, 'pi_89'],
        [99, 'pi_99']
    ]

    for i, (price, col) in enumerate(gg_examples, start=5):
        ws_gg[f'A{i}'] = price
        ws_gg[f'B{i}'] = col
        ws_gg[f'A{i}'].fill = GRAY_EXAMPLE
        ws_gg[f'B{i}'].fill = GRAY_EXAMPLE

    ws_gg.column_dimensions['A'].width = 15
    ws_gg.column_dimensions['B'].width = 30

    wb.save(output_file)
    print(f"Created: {output_file}")


def create_keydriver_annotated(output_file):
    """Create KeyDriver_Config_Template_Annotated.xlsx"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ===== Instructions Sheet =====
    ws_inst = wb.create_sheet("Instructions", 0)
    ws_inst['A1'] = 'TURAS Key Driver Analysis - Configuration Template'
    ws_inst['A1'].font = Font(bold=True, size=14, color='4472C4')

    instructions = [
        '',
        'This template configures key driver analysis to identify which variables drive your outcome metric.',
        '',
        'SHEETS IN THIS TEMPLATE:',
        '  • Settings - Basic analysis parameters',
        '  • Variables - Define outcome and driver variables',
        '',
        'KEY DRIVER ANALYSIS:',
        '  • Identifies which independent variables (drivers) most influence your outcome',
        '  • Uses regression-based relative importance analysis',
        '  • Reports standardized coefficients and relative weights',
        '',
        'QUICK START:',
        '  1. Update Settings sheet with your data file path',
        '  2. In Variables sheet, set Type="Outcome" for your dependent variable (1 variable)',
        '  3. Set Type="Driver" for all independent variables',
        '  4. Optionally set Type="Weight" for survey weight variable',
        '  5. Save and run: turas_load("keydriver") then run analysis',
        '',
        'DOCUMENTATION:',
        '  • Yellow rows = parameter documentation (Required?, Valid Values, Description)',
        '  • Gray rows = example values',
        '  • Delete documentation rows before production use (optional)',
        '',
        'For detailed help, see: modules/keydriver/documentation/'
    ]

    for i, text in enumerate(instructions, start=2):
        ws_inst[f'A{i}'] = text

    ws_inst.column_dimensions['A'].width = 90

    # ===== Settings Sheet =====
    ws_settings = wb.create_sheet("Settings")

    ws_settings['A1'] = 'Setting'
    ws_settings['B1'] = 'Value'
    for cell in ['A1', 'B1']:
        ws_settings[cell].font = BOLD_WHITE_FONT
        ws_settings[cell].fill = BLUE_HEADER

    settings_params = [
        ['analysis_name', 'Optional', 'Any text', 'Descriptive name for analysis (appears in output)', 'Brand Health Drivers'],
        ['data_file', 'Required', 'CSV, XLSX, SAV, RDS file path', 'Path to data file (relative or absolute)', 'keydriver_data.csv'],
        ['output_file', 'Required', 'XLSX file path', 'Output Excel file path', 'keydriver_results.xlsx'],
        ['min_sample_size', 'Optional', 'Positive integer', 'Minimum sample size for analysis (default: 30)', '30']
    ]

    row = 2
    for param, required, valid, desc, example in settings_params:
        ws_settings[f'A{row}'] = param
        ws_settings[f'A{row}'].font = BOLD_BLACK_FONT
        row += 1
        ws_settings[f'A{row}'] = required
        ws_settings[f'B{row}'] = valid
        ws_settings[f'A{row}'].fill = YELLOW_DOC
        ws_settings[f'B{row}'].fill = YELLOW_DOC
        ws_settings[f'A{row}'].font = Font(italic=True, size=9)
        ws_settings[f'B{row}'].font = Font(italic=True, size=9)
        row += 1
        ws_settings[f'A{row}'] = 'Description'
        ws_settings.merge_cells(f'B{row}:B{row}')
        ws_settings[f'B{row}'] = desc
        ws_settings[f'A{row}'].fill = YELLOW_DOC
        ws_settings[f'B{row}'].fill = YELLOW_DOC
        ws_settings[f'A{row}'].font = Font(italic=True, size=9)
        ws_settings[f'B{row}'].font = Font(italic=True, size=9)
        ws_settings[f'B{row}'].alignment = Alignment(wrap_text=True)
        row += 1
        ws_settings[f'A{row}'] = param
        ws_settings[f'B{row}'] = example
        ws_settings[f'A{row}'].fill = GRAY_EXAMPLE
        ws_settings[f'B{row}'].fill = GRAY_EXAMPLE
        row += 1
        row += 1

    ws_settings.column_dimensions['A'].width = 25
    ws_settings.column_dimensions['B'].width = 50

    # ===== Variables Sheet =====
    ws_vars = wb.create_sheet("Variables")

    ws_vars['A1'] = 'VariableName'
    ws_vars['B1'] = 'Type'
    ws_vars['C1'] = 'Label'
    for cell in ['A1', 'B1', 'C1']:
        ws_vars[cell].font = BOLD_WHITE_FONT
        ws_vars[cell].fill = BLUE_HEADER

    # Documentation rows
    ws_vars['A2'] = 'Required?'
    ws_vars['B2'] = 'Valid Values'
    ws_vars['C2'] = 'Description'
    for cell in ['A2', 'B2', 'C2']:
        ws_vars[cell].fill = YELLOW_DOC
        ws_vars[cell].font = Font(italic=True, size=9)

    ws_vars['A3'] = 'Required (column name)'
    ws_vars['B3'] = 'Outcome | Driver | Weight'
    ws_vars['C3'] = 'Display name for output'
    for cell in ['A3', 'B3', 'C3']:
        ws_vars[cell].fill = YELLOW_DOC
        ws_vars[cell].font = Font(italic=True, size=9)

    ws_vars['A4'] = 'Description'
    ws_vars.merge_cells('B4:C4')
    ws_vars['B4'] = 'Variable name must match column in data. Set one Outcome (dependent variable), multiple Drivers (predictors), optional Weight variable.'
    ws_vars['A4'].fill = YELLOW_DOC
    ws_vars['B4'].fill = YELLOW_DOC
    ws_vars['A4'].font = Font(italic=True, size=9)
    ws_vars['B4'].font = Font(italic=True, size=9)
    ws_vars['B4'].alignment = Alignment(wrap_text=True)

    # Example data
    vars_examples = [
        ['overall_satisfaction', 'Outcome', 'Overall Satisfaction'],
        ['product_quality', 'Driver', 'Product Quality'],
        ['customer_service', 'Driver', 'Customer Service'],
        ['value_for_money', 'Driver', 'Value for Money'],
        ['brand_reputation', 'Driver', 'Brand Reputation']
    ]

    for i, (var, type_, label) in enumerate(vars_examples, start=5):
        ws_vars[f'A{i}'] = var
        ws_vars[f'B{i}'] = type_
        ws_vars[f'C{i}'] = label
        for cell in [f'A{i}', f'B{i}', f'C{i}']:
            ws_vars[cell].fill = GRAY_EXAMPLE

    ws_vars.column_dimensions['A'].width = 25
    ws_vars.column_dimensions['B'].width = 12
    ws_vars.column_dimensions['C'].width = 30

    wb.save(output_file)
    print(f"Created: {output_file}")


def create_conjoint_annotated(output_file):
    """Create Conjoint_Config_Template_Annotated.xlsx"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ===== Instructions Sheet =====
    ws_inst = wb.create_sheet("Instructions", 0)
    ws_inst['A1'] = 'TURAS Conjoint Analysis - Configuration Template'
    ws_inst['A1'].font = Font(bold=True, size=14, color='4472C4')

    instructions = [
        '',
        'This template configures choice-based conjoint analysis to estimate attribute part-worth utilities.',
        '',
        'SHEETS IN THIS TEMPLATE:',
        '  • Settings - Analysis type and data file mapping',
        '  • Attributes - Product attributes and their levels',
        '',
        'CONJOINT ANALYSIS:',
        '  • Estimates part-worth utilities for each attribute level',
        '  • Requires choice-based data (respondents choosing from alternatives)',
        '  • Calculates relative importance of each attribute',
        '  • Simulates market share for different product configurations',
        '',
        'DATA FORMAT:',
        '  • One row per alternative shown to respondent',
        '  • choice_set_id: Groups alternatives shown together',
        '  • chosen: 1 if selected, 0 if not',
        '  • respondent_id: Unique respondent identifier',
        '  • Attribute columns: Values for each attribute (matching LevelNames)',
        '',
        'QUICK START:',
        '  1. Update Settings sheet with your data file and column names',
        '  2. Define your attributes in Attributes sheet',
        '  3. Ensure data format matches (one row per alternative)',
        '  4. Save and run: turas_load("conjoint") then run analysis',
        '',
        'DOCUMENTATION:',
        '  • Yellow rows = parameter documentation (Required?, Valid Values, Description)',
        '  • Gray rows = example values',
        '  • Delete documentation rows before production use (optional)',
        '',
        'For detailed help, see: modules/conjoint/documentation/'
    ]

    for i, text in enumerate(instructions, start=2):
        ws_inst[f'A{i}'] = text

    ws_inst.column_dimensions['A'].width = 90

    # ===== Settings Sheet =====
    ws_settings = wb.create_sheet("Settings")

    ws_settings['A1'] = 'Setting'
    ws_settings['B1'] = 'Value'
    for cell in ['A1', 'B1']:
        ws_settings[cell].font = BOLD_WHITE_FONT
        ws_settings[cell].fill = BLUE_HEADER

    settings_params = [
        ['analysis_type', 'Required', 'choice', 'Analysis type (currently only "choice" is supported)', 'choice'],
        ['data_file', 'Required', 'CSV, XLSX, SAV, RDS file path', 'Path to conjoint data file', 'conjoint_data.csv'],
        ['output_file', 'Required', 'XLSX file path', 'Output Excel file path', 'conjoint_results.xlsx'],
        ['choice_set_column', 'Required', 'Column name', 'Column identifying which alternatives were shown together', 'choice_set_id'],
        ['chosen_column', 'Required', 'Column name', 'Column with 1/0 indicating if alternative was chosen', 'chosen'],
        ['respondent_id_column', 'Required', 'Column name', 'Column with unique respondent identifier', 'resp_id']
    ]

    row = 2
    for param, required, valid, desc, example in settings_params:
        ws_settings[f'A{row}'] = param
        ws_settings[f'A{row}'].font = BOLD_BLACK_FONT
        row += 1
        ws_settings[f'A{row}'] = required
        ws_settings[f'B{row}'] = valid
        ws_settings[f'A{row}'].fill = YELLOW_DOC
        ws_settings[f'B{row}'].fill = YELLOW_DOC
        ws_settings[f'A{row}'].font = Font(italic=True, size=9)
        ws_settings[f'B{row}'].font = Font(italic=True, size=9)
        row += 1
        ws_settings[f'A{row}'] = 'Description'
        ws_settings.merge_cells(f'B{row}:B{row}')
        ws_settings[f'B{row}'] = desc
        ws_settings[f'A{row}'].fill = YELLOW_DOC
        ws_settings[f'B{row}'].fill = YELLOW_DOC
        ws_settings[f'A{row}'].font = Font(italic=True, size=9)
        ws_settings[f'B{row}'].font = Font(italic=True, size=9)
        ws_settings[f'B{row}'].alignment = Alignment(wrap_text=True)
        row += 1
        ws_settings[f'A{row}'] = param
        ws_settings[f'B{row}'] = example
        ws_settings[f'A{row}'].fill = GRAY_EXAMPLE
        ws_settings[f'B{row}'].fill = GRAY_EXAMPLE
        row += 1
        row += 1

    ws_settings.column_dimensions['A'].width = 25
    ws_settings.column_dimensions['B'].width = 50

    # ===== Attributes Sheet =====
    ws_attrs = wb.create_sheet("Attributes")

    ws_attrs['A1'] = 'AttributeName'
    ws_attrs['B1'] = 'NumLevels'
    ws_attrs['C1'] = 'LevelNames'
    for cell in ['A1', 'B1', 'C1']:
        ws_attrs[cell].font = BOLD_WHITE_FONT
        ws_attrs[cell].fill = BLUE_HEADER

    # Documentation rows
    ws_attrs['A2'] = 'Required?'
    ws_attrs['B2'] = 'Valid Values'
    ws_attrs['C2'] = 'Description'
    for cell in ['A2', 'B2', 'C2']:
        ws_attrs[cell].fill = YELLOW_DOC
        ws_attrs[cell].font = Font(italic=True, size=9)

    ws_attrs['A3'] = 'Required (column name)'
    ws_attrs['B3'] = 'Integer (2+)'
    ws_attrs['C3'] = 'Comma-separated list'
    for cell in ['A3', 'B3', 'C3']:
        ws_attrs[cell].fill = YELLOW_DOC
        ws_attrs[cell].font = Font(italic=True, size=9)

    ws_attrs['A4'] = 'Description'
    ws_attrs.merge_cells('B4:C4')
    ws_attrs['B4'] = 'Attribute name must match data column. NumLevels is count of levels. LevelNames are comma-separated values matching data.'
    ws_attrs['A4'].fill = YELLOW_DOC
    ws_attrs['B4'].fill = YELLOW_DOC
    ws_attrs['A4'].font = Font(italic=True, size=9)
    ws_attrs['B4'].font = Font(italic=True, size=9)
    ws_attrs['B4'].alignment = Alignment(wrap_text=True)

    # Example data
    attrs_examples = [
        ['Price', 3, '£449, £599, £699'],
        ['Brand', 3, 'Apple, Samsung, Google'],
        ['Storage', 3, '128GB, 256GB, 512GB'],
        ['Battery', 3, '12 hours, 18 hours, 24 hours']
    ]

    for i, (attr, num, levels) in enumerate(attrs_examples, start=5):
        ws_attrs[f'A{i}'] = attr
        ws_attrs[f'B{i}'] = num
        ws_attrs[f'C{i}'] = levels
        for cell in [f'A{i}', f'B{i}', f'C{i}']:
            ws_attrs[cell].fill = GRAY_EXAMPLE

    ws_attrs.column_dimensions['A'].width = 18
    ws_attrs.column_dimensions['B'].width = 12
    ws_attrs.column_dimensions['C'].width = 45

    wb.save(output_file)
    print(f"Created: {output_file}")


if __name__ == "__main__":
    import os

    # Create templates in /templates directory
    templates_dir = "/home/user/Turas/templates"

    create_pricing_annotated(os.path.join(templates_dir, "Pricing_Config_Template_Annotated.xlsx"))
    create_keydriver_annotated(os.path.join(templates_dir, "KeyDriver_Config_Template_Annotated.xlsx"))
    create_conjoint_annotated(os.path.join(templates_dir, "Conjoint_Config_Template_Annotated.xlsx"))

    print("\nAll annotated templates created successfully!")
