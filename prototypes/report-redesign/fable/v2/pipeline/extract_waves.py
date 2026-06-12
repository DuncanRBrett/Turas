#!/usr/bin/env python3
"""Extract results from SACAP crosstabs workbooks (2018-2024) into one
multi-wave payload for the v2 report's tracking data layer — Total column
AND every banner-segment column that maps onto a 2025 banner column.

Question codes shift between waves, so downstream matching is by normalised
question title (occurrence-ordered for duplicate titles) + row label.
Segments are matched by normalised column label (plus segment_aliases).

Row taxonomy per question:
  - "Base (n=)"                      -> question base (Total + per segment)
  - label + Frequency / Column %     -> category row {label, n, pct, seg{}}
  - label + Column % (no Frequency)  -> NET-style summary row {label, pct, seg{}}
    (the old single-wave extractor dropped these -- NETs, NET POSITIVE and
    Detractor/Passive/Promoter rows live here)
  - Index / Average / Score rows     -> stats {index|mean|nps} + seg_stats

Per-segment cells store Column % only; counts are recovered downstream as
round(pct/100 * segment base), matching the published-totals sig convention.
Segment coverage varies by year (2018/2019/2023 are Total-only; 2020-2022
carry Campus; 2024 carries Campus + Intensity + Course) — consumers must
treat segment data as sparse.

Question titles were reworded over the years; wave_title_aliases.json maps
historical normalised titles onto their 2025 equivalents (the tracker
module's question-mapper pattern, kept reviewable as data). The same file's
"segment_aliases" maps historical column labels (e.g. "Online") onto 2025
banner labels ("Online campus").

Usage:
  extract_waves.py [--aliases <aliases.json>] <agg_2025.json> <out.json> \
      <year>=<crosstabs.xlsx> ...
"""

import json
import re
import sys

import openpyxl


def norm(text):
    """Normalise a title/label for cross-wave matching (mirrors TR.model.norm)."""
    t = re.sub(r"\s+", " ", str(text or "")).strip().lower()
    t = re.sub(r"[^a-z0-9 ]", "", t)
    return t


STAT_KEYS = {"Index": "index", "Average": "mean", "Score": "nps"}


def find_segments(ws, agg_col_norms, seg_aliases):
    """Locate the header row and map workbook columns onto 2025 banner
    columns. Returns (segments, col_map) where segments is the ordered
    [{label, norm}] list (canonical 2025 labels) and col_map maps workbook
    column index -> canonical norm. Empty when the wave is Total-only."""
    header = None
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        if len(cells) >= 3 and cells[0] == "Question" and cells[2] == "Total":
            header = cells
            break
    if header is None:
        return [], {}
    segments, col_map, seen = [], {}, set()
    for idx in range(3, len(header)):
        label = header[idx]
        if not label:
            continue
        n = norm(label)
        n = seg_aliases.get(n, n)
        if n not in agg_col_norms or n in seen:
            continue
        seen.add(n)
        segments.append({"label": agg_col_norms[n], "norm": n})
        col_map[idx] = n
    return segments, col_map


def seg_values(row, col_map):
    """Per-segment numeric values of one sheet row ({} when none parse)."""
    out = {}
    for idx, seg in col_map.items():
        v = row[idx] if idx < len(row) else None
        if v is None or v == "":
            continue
        try:
            out[seg] = float(v)
        except (TypeError, ValueError):
            continue
    return out


def parse_workbook(path, agg_col_norms, seg_aliases):
    """Parse one crosstabs workbook into (ordered question list, segments)."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Crosstabs"]
    # the workbooks ship broken dimension records (A1:A1); recalculate
    ws.reset_dimensions()
    segments, col_map = find_segments(ws, agg_col_norms, seg_aliases)
    max_col = max(col_map.keys(), default=3) + 1

    questions = []
    current = None
    pending_label = None

    for row in ws.iter_rows(min_col=1, max_col=max_col, values_only=True):
        a = row[0]
        b = row[1] if len(row) > 1 else None
        total = row[2] if len(row) > 2 else None
        if a and isinstance(a, str) and re.match(r"Q\d+\s*-\s*", a):
            code, title = re.match(r"(Q\d+)\s*-\s*(.*)", a).groups()
            current = {"code": code, "title": title.strip(), "base": None,
                       "bases": {}, "rows": {}, "stats": {}, "seg_stats": {}}
            questions.append(current)
            pending_label = None
            continue
        if current is None:
            continue
        if b == "Base (n=)":
            current["base"] = int(float(total)) if total is not None else None
            for seg, v in seg_values(row, col_map).items():
                current["bases"][seg] = int(v)
            continue
        if a and b == "Frequency":
            pending_label = str(a).strip()
            entry = current["rows"].setdefault(norm(pending_label), {})
            entry["label"] = pending_label
            entry["n"] = int(float(total or 0))
            continue
        if b == "Column %":
            seg = seg_values(row, col_map)
            if pending_label:
                # second half of a category row
                entry = current["rows"][norm(pending_label)]
                entry["pct"] = float(total or 0)
                if seg:
                    entry["seg"] = seg
                pending_label = None
            elif a:
                # NET-style summary row: percentage only, no frequency
                label = str(a).strip()
                entry = {"label": label, "pct": float(total or 0)}
                if seg:
                    entry["seg"] = seg
                current["rows"][norm(label)] = entry
            continue
        stat = STAT_KEYS.get(str(b).strip()) if b else None
        if stat and total is not None:
            try:
                current["stats"][stat] = float(total)
            except (TypeError, ValueError):
                continue
            for seg, v in seg_values(row, col_map).items():
                current["seg_stats"].setdefault(seg, {})[stat] = v

    # drop empty per-question segment dicts to keep the payload compact
    for q in questions:
        if not q["bases"]:
            del q["bases"]
        if not q["seg_stats"]:
            del q["seg_stats"]
    return questions, segments


def with_match_keys(questions, aliases):
    """Attach occurrence-ordered match keys (duplicate titles pair k-th to k-th).
    title_norm is the CANONICAL (alias-resolved) norm the JS engine matches on."""
    seen = {}
    for q in questions:
        t = norm(q["title"])
        t = aliases.get(t, t)
        k = seen.get(t, 0)
        seen[t] = k + 1
        q["title_norm"] = t
        q["match_key"] = t if k == 0 else t + "#" + str(k)
    return questions


def match_report(agg_questions, wave_questions, segments):
    """How many 2025 questions find a matching question in this wave."""
    wave_keys = {q["match_key"] for q in wave_questions}
    matched = sum(1 for q in agg_questions if q["match_key"] in wave_keys)
    return {"matched": matched, "of": len(agg_questions),
            "rate": round(matched / len(agg_questions), 3) if agg_questions else 0,
            "segments": len(segments)}


def main():
    args = sys.argv[1:]
    aliases, seg_aliases = {}, {}
    if args and args[0] == "--aliases":
        with open(args[1], encoding="utf-8") as f:
            alias_doc = json.load(f)
        aliases = alias_doc["aliases"]
        seg_aliases = alias_doc.get("segment_aliases", {})
        args = args[2:]
    if len(args) < 3:
        print(__doc__)
        return 1
    with open(args[0], encoding="utf-8") as f:
        agg = json.load(f)
    # 2025 titles are canonical: never alias-resolved themselves
    agg_questions = with_match_keys(
        [{"title": q["title"]} for q in agg["questions"]], {})
    unused = set(aliases) - {norm(a) for a in aliases}
    if unused:
        print(f"REFUSED: alias keys not normalised: {sorted(unused)[:3]}")
        return 1
    bad_targets = {v for v in aliases.values()} - \
        {q["title_norm"] for q in agg_questions}
    if bad_targets:
        print(f"REFUSED: alias targets missing from 2025: {sorted(bad_targets)[:3]}")
        return 1
    # canonical 2025 banner columns: norm -> label (Total excluded)
    agg_col_norms = {norm(c["label"]): c["label"]
                     for c in agg["columns"] if c.get("group") != "total"}
    bad_segs = set(seg_aliases.values()) - set(agg_col_norms)
    if bad_segs:
        print(f"REFUSED: segment alias targets missing from 2025: {sorted(bad_segs)[:3]}")
        return 1

    waves, report = [], {}
    for arg in args[2:]:
        year, _, path = arg.partition("=")
        if not year.isdigit() or not path:
            print(f"REFUSED: bad wave argument {arg!r} (want <year>=<xlsx>)")
            return 1
        questions, segments = parse_workbook(path, agg_col_norms, seg_aliases)
        questions = with_match_keys(questions, aliases)
        waves.append({"wave": f"Annual {year}", "year": int(year),
                      "segments": segments, "questions": questions})
        report[year] = match_report(agg_questions, questions, segments)
        print(f"  {year}: {len(questions)} questions, "
              f"{sum(len(q['rows']) for q in questions)} rows, "
              f"{len(segments)} segments, "
              f"matched {report[year]['matched']}/{report[year]['of']} "
              f"({report[year]['rate']:.0%}) of 2025 questions")

    waves.sort(key=lambda w: w["year"])
    out = {"schema_version": 3, "waves": waves, "match_report": report}
    with open(args[1], "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
    print(f"OK {len(waves)} waves -> {args[1]}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
