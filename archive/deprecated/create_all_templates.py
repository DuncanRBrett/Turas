#!/usr/bin/env python3
"""
Create All Turas Module Templates
==================================
This script creates Excel template files for all Turas modules with:
- Clear headers and example data
- Professional formatting
- Helpful comments and instructions
"""

import os
from datetime import date
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(['pip3', 'install', '--quiet', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Create templates directory
TEMPLATES_DIR = "templates"
if not os.path.exists(TEMPLATES_DIR):
    os.makedirs(TEMPLATES_DIR)

print("Creating Turas module templates...\n")

def create_header_style():
    """Create styled header for Excel sheets"""
    return {
        'font': Font(name='Arial', size=11, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def apply_header_style(ws, row, max_col):
    """Apply header styling to a row"""
    style = create_header_style()
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = style['font']
        cell.fill = style['fill']
        cell.alignment = style['alignment']
        cell.border = style['border']

def set_column_widths(ws, widths):
    """Set column widths"""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

# ==============================================================================
# PARSER MODULE TEMPLATES
# ==============================================================================

print("Creating Parser templates...")

def create_parser_questionnaire_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Questionnaire"

    # Headers
    headers = ["Q_Number", "Question_Text", "Response_Options"]
    ws.append(headers)
    apply_header_style(ws, 1, len(headers))

    # Example data
    data = [
        ["Q1", "Which of the following brands are you aware of? (Select all that apply)",
         "Brand A, Brand B, Brand C, Brand D, Other, None"],
        ["Q2", "Which ONE brand do you prefer?",
         "Brand A, Brand B, Brand C, Brand D"],
        ["Q3", "How satisfied are you with [BRAND]? (1=Very Dissatisfied, 5=Very Satisfied)",
         "1, 2, 3, 4, 5"],
        ["Q4", "How likely are you to recommend [BRAND] to a friend? (0=Not at all likely, 10=Extremely likely)",
         "0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10"],
        ["Q5", "In the last 3 months, how many times have you purchased [PRODUCT]?",
         "Numeric"],
        ["Q6", "What do you like most about [BRAND]? (Open-ended)",
         "Open-ended"]
    ]

    for row in data:
        ws.append(row)

    set_column_widths(ws, [12, 80, 50])

    # Add instructions
    ws.append([])
    ws.append([])
    instructions = [
        ["INSTRUCTIONS:"],
        ["1. Enter your survey questions in this template"],
        ["2. Q_Number: Unique identifier for each question (e.g., Q1, Q2, Q3)"],
        ["3. Question_Text: Full question wording as shown to respondents"],
        ["4. Response_Options: Comma-separated list of response options"],
        ["   - For multi-select: List all options"],
        ["   - For single choice: List all options"],
        ["   - For rating scales: List all scale points (e.g., 1, 2, 3, 4, 5)"],
        ["   - For numeric: Enter 'Numeric'"],
        ["   - For open-ended: Enter 'Open-ended'"],
        ["5. Run Parser to generate Survey_Structure.xlsx"]
    ]

    for instruction in instructions:
        ws.append(instruction)

    filepath = os.path.join(TEMPLATES_DIR, "Parser_Questionnaire_Template.xlsx")
    wb.save(filepath)
    print(f"  ✓ Parser_Questionnaire_Template.xlsx")

create_parser_questionnaire_template()

# ==============================================================================
# TABS MODULE TEMPLATES
# ==============================================================================

print("\nCreating Tabs templates...")

def create_tabs_survey_structure_template():
    wb = Workbook()

    # Questions Sheet
    ws_q = wb.active
    ws_q.title = "Questions"

    headers_q = ["QuestionCode", "QuestionText", "Variable_Type", "Columns"]
    ws_q.append(headers_q)
    apply_header_style(ws_q, 1, len(headers_q))

    data_q = [
        ["Q01", "Brand Awareness (Unaided)", "Single_Response", ""],
        ["Q02", "Brand Consideration", "Single_Response", ""],
        ["Q03", "Brand Preference", "Single_Response", ""],
        ["Q04", "Overall Satisfaction (1-5)", "Rating", ""],
        ["Q05", "Likelihood to Recommend (0-10)", "NPS", ""],
        ["Gender", "Gender", "Single_Response", ""],
        ["Age_Group", "Age Group", "Single_Response", ""]
    ]

    for row in data_q:
        ws_q.append(row)

    set_column_widths(ws_q, [15, 40, 18, 12])

    # Options Sheet
    ws_o = wb.create_sheet("Options")

    headers_o = ["QuestionCode", "OptionValue", "OptionText", "ShowInOutput"]
    ws_o.append(headers_o)
    apply_header_style(ws_o, 1, len(headers_o))

    data_o = [
        # Q01
        ["Q01", 1, "Brand A", "TRUE"],
        ["Q01", 2, "Brand B", "TRUE"],
        ["Q01", 3, "Brand C", "TRUE"],
        ["Q01", 4, "None", "TRUE"],
        # Q02
        ["Q02", 1, "Brand A", "TRUE"],
        ["Q02", 2, "Brand B", "TRUE"],
        ["Q02", 3, "Brand C", "TRUE"],
        # Q03
        ["Q03", 1, "Brand A", "TRUE"],
        ["Q03", 2, "Brand B", "TRUE"],
        ["Q03", 3, "Brand C", "TRUE"],
        # Q04
        ["Q04", 1, "1 - Very Dissatisfied", "TRUE"],
        ["Q04", 2, "2", "TRUE"],
        ["Q04", 3, "3", "TRUE"],
        ["Q04", 4, "4", "TRUE"],
        ["Q04", 5, "5 - Very Satisfied", "TRUE"],
        # Q05 (NPS)
        ["Q05", 0, "0", "TRUE"],
        ["Q05", 1, "1", "TRUE"],
        ["Q05", 2, "2", "TRUE"],
        ["Q05", 3, "3", "TRUE"],
        ["Q05", 4, "4", "TRUE"],
        ["Q05", 5, "5", "TRUE"],
        ["Q05", 6, "6", "TRUE"],
        ["Q05", 7, "7", "TRUE"],
        ["Q05", 8, "8", "TRUE"],
        ["Q05", 9, "9", "TRUE"],
        ["Q05", 10, "10", "TRUE"],
        # Gender
        ["Gender", 1, "Male", "TRUE"],
        ["Gender", 2, "Female", "TRUE"],
        # Age_Group
        ["Age_Group", 1, "18-34", "TRUE"],
        ["Age_Group", 2, "35-54", "TRUE"],
        ["Age_Group", 3, "55+", "TRUE"]
    ]

    for row in data_o:
        ws_o.append(row)

    set_column_widths(ws_o, [15, 12, 30, 12])

    filepath = os.path.join(TEMPLATES_DIR, "Tabs_Survey_Structure_Template.xlsx")
    wb.save(filepath)
    print(f"  ✓ Tabs_Survey_Structure_Template.xlsx")

def create_tabs_config_template():
    wb = Workbook()

    # Settings Sheet
    ws_settings = wb.active
    ws_settings.title = "Settings"

    headers = ["Setting", "Value", "Description"]
    ws_settings.append(headers)
    apply_header_style(ws_settings, 1, len(headers))

    data = [
        ["survey_structure_file", "Survey_Structure.xlsx", "Path to Survey_Structure.xlsx file"],
        ["data_file", "survey_data.csv", "Path to survey data file (CSV, XLSX, SAV, DTA)"],
        ["output_file", "Crosstab_Results.xlsx", "Path to output Excel file"],
        ["show_significance", "TRUE", "Show significance testing (TRUE/FALSE)"],
        ["significance_level", "0.05", "Significance level (0.05 = 95% confidence, 0.10 = 90%)"],
        ["minimum_base", "30", "Minimum base size for significance testing"],
        ["stat_test", "chi-square", "Statistical test: chi-square, z-test, or t-test"],
        ["decimal_places", "0", "Decimal places for percentages (0 = whole numbers)"],
        ["decimal_places_average", "1", "Decimal places for averages"],
        ["show_frequencies", "TRUE", "Show frequency counts (TRUE/FALSE)"],
        ["show_percentages", "TRUE", "Show column percentages (TRUE/FALSE)"],
        ["weight_column", "NA", "Weight column name (or NA if unweighted)"]
    ]

    for row in data:
        ws_settings.append(row)

    set_column_widths(ws_settings, [25, 25, 55])

    # Banner Sheet
    ws_banner = wb.create_sheet("Banner")
    ws_banner.append(["BannerQuestion"])
    apply_header_style(ws_banner, 1, 1)

    banner_data = [["Total"], ["Gender"], ["Age_Group"]]
    for row in banner_data:
        ws_banner.append(row)

    set_column_widths(ws_banner, [20])

    # Stub Sheet
    ws_stub = wb.create_sheet("Stub")
    headers_stub = ["StubQuestion", "BaseFilter"]
    ws_stub.append(headers_stub)
    apply_header_style(ws_stub, 1, len(headers_stub))

    stub_data = [
        ["Q01", ""],
        ["Q02", ""],
        ["Q03", ""],
        ["Q04", "Gender == 'Male'"],
        ["Q05", ""]
    ]

    for row in stub_data:
        ws_stub.append(row)

    set_column_widths(ws_stub, [20, 30])

    filepath = os.path.join(TEMPLATES_DIR, "Tabs_Config_Template.xlsx")
    wb.save(filepath)
    print(f"  ✓ Tabs_Config_Template.xlsx")

create_tabs_survey_structure_template()
create_tabs_config_template()

# ==============================================================================
# TRACKER MODULE TEMPLATES
# ==============================================================================

print("\nCreating Tracker templates...")

def create_tracker_config_template():
    wb = Workbook()

    # Waves Sheet
    ws_waves = wb.active
    ws_waves.title = "Waves"

    headers = ["WaveID", "WaveName", "DataFile", "FieldworkStart", "FieldworkEnd", "WeightVariable"]
    ws_waves.append(headers)
    apply_header_style(ws_waves, 1, len(headers))

    data = [
        ["W1", "Q1 2024", "wave1.csv", "2024-01-01", "2024-01-15", "Weight"],
        ["W2", "Q2 2024", "wave2.csv", "2024-04-01", "2024-04-15", "Weight"],
        ["W3", "Q3 2024", "wave3.csv", "2024-07-01", "2024-07-15", "Weight"],
        ["W4", "Q4 2024", "wave4.csv", "2024-10-01", "2024-10-15", "Weight"]
    ]

    for row in data:
        ws_waves.append(row)

    set_column_widths(ws_waves, [10, 15, 20, 15, 15, 15])

    # TrackedQuestions Sheet
    ws_questions = wb.create_sheet("TrackedQuestions")

    headers_q = ["QuestionCode", "QuestionText", "QuestionType"]
    ws_questions.append(headers_q)
    apply_header_style(ws_questions, 1, len(headers_q))

    data_q = [
        ["Q01_Awareness", "Brand Awareness (Unaided)", "proportion"],
        ["Q02_Consideration", "Brand Consideration", "proportion"],
        ["Q03_Preference", "Brand Preference", "proportion"],
        ["Q04_Satisfaction", "Overall Satisfaction (1-5)", "rating"],
        ["Q05_NPS", "Net Promoter Score (0-10)", "nps"]
    ]

    for row in data_q:
        ws_questions.append(row)

    set_column_widths(ws_questions, [20, 40, 15])

    # Banner Sheet
    ws_banner = wb.create_sheet("Banner")

    headers_b = ["BreakVariable", "BreakLabel"]
    ws_banner.append(headers_b)
    apply_header_style(ws_banner, 1, len(headers_b))

    data_b = [
        ["Total", "Total"],
        ["Gender", "Gender"],
        ["Age_Group", "Age Group"]
    ]

    for row in data_b:
        ws_banner.append(row)

    set_column_widths(ws_banner, [20, 20])

    # Settings Sheet
    ws_settings = wb.create_sheet("Settings")

    headers_s = ["SettingName", "SettingValue"]
    ws_settings.append(headers_s)
    apply_header_style(ws_settings, 1, len(headers_s))

    data_s = [
        ["project_name", "2024 Brand Tracking Study"],
        ["output_file", "Tracking_Results.xlsx"],
        ["confidence_level", "0.95"],
        ["min_base_size", "30"],
        ["trend_significance", "TRUE"],
        ["decimal_places_proportion", "0"],
        ["decimal_places_mean", "2"],
        ["show_sample_sizes", "TRUE"]
    ]

    for row in data_s:
        ws_settings.append(row)

    set_column_widths(ws_settings, [30, 30])

    filepath = os.path.join(TEMPLATES_DIR, "Tracker_Config_Template.xlsx")
    wb.save(filepath)
    print(f"  ✓ Tracker_Config_Template.xlsx")

def create_tracker_question_mapping_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "QuestionMap"

    headers = ["QuestionCode", "QuestionText", "QuestionType", "W1", "W2", "W3", "W4"]
    ws.append(headers)
    apply_header_style(ws, 1, len(headers))

    data = [
        ["Q01_Awareness", "Brand Awareness (Unaided)", "proportion", "Q1_Awareness", "Q1_Awareness", "Q01_Aware", "Q01_Aware"],
        ["Q02_Consideration", "Brand Consideration", "proportion", "Q2_Consider", "Q2_Consider", "Q02_Consideration", "Q02_Consideration"],
        ["Q03_Preference", "Brand Preference", "proportion", "Q3_Preference", "Q3_Preference", "Q03_Pref", "Q03_Pref"],
        ["Q04_Satisfaction", "Overall Satisfaction (1-5)", "rating", "Q4_Sat", "Q4_Sat", "Q04_Satisfaction", "Q04_Satisfaction"],
        ["Q05_NPS", "Net Promoter Score (0-10)", "nps", "Q5_NPS", "Q5_NPS", "Q05_NPS_Score", "Q05_NPS_Score"]
    ]

    for row in data:
        ws.append(row)

    set_column_widths(ws, [20, 40, 15, 18, 18, 20, 20])

    # Add instructions
    ws.append([])
    ws.append(["INSTRUCTIONS:"])
    ws.append(["- QuestionCode: Standardized question identifier (used in Tracking_Config.xlsx)"])
    ws.append(["- QuestionText: Question wording"])
    ws.append(["- QuestionType: proportion, rating, nps, or composite"])
    ws.append(["- W1, W2, W3, W4: Actual column names in each wave's data file"])
    ws.append(["- Use NA if question not asked in that wave"])
    ws.append(["- Add more wave columns (W5, W6, etc.) as needed"])

    filepath = os.path.join(TEMPLATES_DIR, "Tracker_Question_Mapping_Template.xlsx")
    wb.save(filepath)
    print(f"  ✓ Tracker_Question_Mapping_Template.xlsx")

create_tracker_config_template()
create_tracker_question_mapping_template()

# ==============================================================================
# CONFIDENCE MODULE TEMPLATES
# ==============================================================================

print("\nCreating Confidence templates...")

def create_confidence_config_template():
    wb = Workbook()

    # Settings Sheet
    ws_settings = wb.active
    ws_settings.title = "Settings"

    headers = ["Setting", "Value", "Description"]
    ws_settings.append(headers)
    apply_header_style(ws_settings, 1, len(headers))

    data = [
        ["data_file", "survey_data.csv", "Path to survey data file"],
        ["survey_structure_file", "Survey_Structure.xlsx", "Path to Survey_Structure.xlsx (from Tabs or Parser)"],
        ["output_file", "Confidence_Analysis.xlsx", "Output Excel file name"],
        ["weight_variable", "Weight", "Weight column name (or NA if unweighted)"],
        ["confidence_level", "0.95", "Confidence level (0.90, 0.95, or 0.99)"],
        ["decimal_separator", ".", "Decimal separator: period (.) or comma (,)"],
        ["bootstrap_iterations", "5000", "Number of bootstrap iterations (1000-10000)"],
        ["methods_proportion", "MOE,Wilson,Bootstrap,Bayesian", "Methods for proportions: MOE, Wilson, Bootstrap, Bayesian"],
        ["methods_mean", "tdist,Bootstrap,Bayesian", "Methods for means: tdist, Bootstrap, Bayesian"]
    ]

    for row in data:
        ws_settings.append(row)

    set_column_widths(ws_settings, [25, 35, 50])

    # Questions Sheet
    ws_questions = wb.create_sheet("Questions")

    headers_q = ["QuestionCode", "QuestionType", "BayesianPrior_Mean", "BayesianPrior_N"]
    ws_questions.append(headers_q)
    apply_header_style(ws_questions, 1, len(headers_q))

    data_q = [
        ["Q01", "proportion", "0.5", "30"],
        ["Q02", "proportion", "0.5", "30"],
        ["Q03", "rating", "3.5", "30"],
        ["Q04", "nps", "25", "30"]
    ]

    for row in data_q:
        ws_questions.append(row)

    set_column_widths(ws_questions, [18, 18, 20, 18])

    # Add instructions
    ws_questions.append([])
    ws_questions.append(["INSTRUCTIONS:"])
    ws_questions.append(["- QuestionCode: Must match codes in Survey_Structure.xlsx"])
    ws_questions.append(["- QuestionType: proportion, rating, or nps"])
    ws_questions.append(["- BayesianPrior_Mean: Prior estimate (e.g., 0.5 = 50% for proportions, 3.5 for 1-5 rating)"])
    ws_questions.append(["- BayesianPrior_N: Prior sample size (strength of prior, typically 30-100)"])
    ws_questions.append(["- Leave Bayesian columns empty if not using Bayesian method"])

    filepath = os.path.join(TEMPLATES_DIR, "Confidence_Config_Template.xlsx")
    wb.save(filepath)
    print(f"  ✓ Confidence_Config_Template.xlsx")

create_confidence_config_template()

# ==============================================================================
# SEGMENT MODULE TEMPLATES
# ==============================================================================

print("\nCreating Segment templates...")

def create_segment_config_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Config"

    headers = ["parameter", "value", "description"]
    ws.append(headers)
    apply_header_style(ws, 1, len(headers))

    data = [
        ["data_file", "survey_data.csv", "Path to survey data file (CSV or XLSX)"],
        ["id_variable", "RespondentID", "Column name containing respondent ID"],
        ["clustering_vars", "Q1,Q2,Q3,Q4,Q5", "Comma-separated list of variables for clustering"],
        ["profiling_vars", "Age,Gender,Region", "Comma-separated list of variables for profiling segments"],
        ["k_min", "3", "Minimum number of segments to test (exploration mode)"],
        ["k_max", "6", "Maximum number of segments to test (exploration mode)"],
        ["k_final", "4", "Final number of segments (final run mode)"],
        ["output_folder", "output/", "Output directory for results"],
        ["random_seed", "123", "Random seed for reproducibility"],
        ["max_iterations", "100", "Maximum k-means iterations"],
        ["n_starts", "25", "Number of random starts for k-means"],
        ["outlier_method", "zscore", "Outlier detection method: zscore or mahalanobis"],
        ["outlier_threshold", "3", "Outlier threshold (z-score units or chi-square critical value)"],
        ["handle_outliers", "remove", "How to handle outliers: remove, flag, or ignore"],
        ["run_mode", "explore", "Run mode: explore (test k_min to k_max) or final (use k_final)"]
    ]

    for row in data:
        ws.append(row)

    set_column_widths(ws, [22, 30, 60])

    filepath = os.path.join(TEMPLATES_DIR, "Segment_Config_Template.xlsx")
    wb.save(filepath)
    print(f"  ✓ Segment_Config_Template.xlsx")

create_segment_config_template()

# ==============================================================================
# SUMMARY
# ==============================================================================

print("\n")
print("=" * 80)
print("TEMPLATE CREATION COMPLETE")
print("=" * 80)
print(f"\nAll templates created in: {os.path.abspath(TEMPLATES_DIR)}\n")
print("Templates created:")
print("  Parser:")
print("    - Parser_Questionnaire_Template.xlsx")
print("  Tabs:")
print("    - Tabs_Survey_Structure_Template.xlsx")
print("    - Tabs_Config_Template.xlsx")
print("  Tracker:")
print("    - Tracker_Config_Template.xlsx")
print("    - Tracker_Question_Mapping_Template.xlsx")
print("  Confidence:")
print("    - Confidence_Config_Template.xlsx")
print("  Segment:")
print("    - Segment_Config_Template.xlsx")
print("\nTotal: 7 template files\n")
