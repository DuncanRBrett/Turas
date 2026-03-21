#!/usr/bin/env python3
"""
Update example config to enable market simulator
"""

from openpyxl import load_workbook

# Load the existing workbook
wb = load_workbook("/home/user/Turas/modules/conjoint/examples/example_config.xlsx")

# Get the Settings sheet
ws = wb["Settings"]

# Update the generate_market_simulator setting
# Find the row with this setting
for row in range(2, 20):  # Check first 20 rows
    cell_value = ws.cell(row=row, column=1).value
    if cell_value == "generate_market_simulator":
        # Update value to TRUE
        ws.cell(row=row, column=2).value = "TRUE"
        print(f"✓ Updated generate_market_simulator to TRUE (row {row})")
        break

# Save the updated workbook
wb.save("/home/user/Turas/modules/conjoint/examples/example_config.xlsx")
print("✓ Config file updated with market simulator enabled")
