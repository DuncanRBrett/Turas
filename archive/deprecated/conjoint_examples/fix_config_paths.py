#!/usr/bin/env python3
"""
Fix configuration file paths - remove duplicate 'examples/' prefix
Since the config file is in modules/conjoint/examples/, paths should be relative to that directory.
"""

import openpyxl
from pathlib import Path

# File path
config_file = Path(__file__).parent / "example_config.xlsx"

print(f"Fixing configuration paths in: {config_file}")

# Load workbook
wb = openpyxl.load_workbook(config_file)

# Fix Settings sheet
if 'Settings' in wb.sheetnames:
    ws = wb['Settings']

    # Find and fix data_file and output_file paths
    for row in ws.iter_rows():
        if row[0].value and isinstance(row[0].value, str):
            setting = row[0].value.strip()

            if setting == 'data_file':
                old_value = row[1].value
                # Change from "examples/sample_cbc_data.csv" to "sample_cbc_data.csv"
                new_value = "sample_cbc_data.csv"
                row[1].value = new_value
                print(f"✓ Fixed data_file: '{old_value}' -> '{new_value}'")

            elif setting == 'output_file':
                old_value = row[1].value
                # Change from "examples/output/example_results.xlsx" to "output/example_results.xlsx"
                new_value = "output/example_results.xlsx"
                row[1].value = new_value
                print(f"✓ Fixed output_file: '{old_value}' -> '{new_value}'")

# Save
wb.save(config_file)
print(f"\n✓ Configuration file updated successfully!")
print(f"\nPaths are now relative to the config file location:")
print(f"  data_file: sample_cbc_data.csv")
print(f"  output_file: output/example_results.xlsx")
