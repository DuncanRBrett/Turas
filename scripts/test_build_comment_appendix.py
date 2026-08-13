#!/usr/bin/env python3
"""
Known-answer tests for scripts/build_comment_appendix.py.

Self-contained (no pytest needed):  python3 scripts/test_build_comment_appendix.py
Covers column detection, the non-destructive incremental update (preservation +
idempotency), id-header handling (ID vs ResponseID), and the empty-column guard.
"""

import importlib.util
import tempfile
from pathlib import Path

import openpyxl
import pandas as pd

_spec = importlib.util.spec_from_file_location(
    "bca", str(Path(__file__).with_name("build_comment_appendix.py")))
bca = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(bca)

_passed = _failed = 0


def check(cond, msg):
    global _passed, _failed
    if cond:
        _passed += 1
    else:
        _failed += 1
        print("  FAIL:", msg)


def sample_df():
    # id + two clearly-named comments, one free-text-but-oddly-named (SACS-style),
    # one closed categorical (few distinct values, repeated -> low uniqueness),
    # one metadata column, one empty comment column.
    return pd.DataFrame({
        "Response ID": [8, 11, 32, 40, 55, 61, 70, 88],
        "Q1Comment": ["the delivery service is consistently excellent", "",
                      "they arrive late far too often these days", "",
                      "helpful and professional staff on every visit", "", "", "quick and reliable every time"],
        "xtraNotes": ["they really do help a lot honestly", "no complaints from us at all here",
                      "", "the range could be a little wider", "", "", "", ""],
        "Please share your view": ["prices have climbed a lot recently", "",
                                   "signage in the store is quite limited", "", "",
                                   "everything is working really well for us", "", ""],
        "Q2": ["Very interested", "Not interested", "Interested", "Very interested",
               "Not interested", "Interested", "Very interested", "Not interested"],
        "Status": ["Complete"] * 8,
        "Q9Comment": [""] * 8,
    })


def sheet_rows(ws):
    hr = bca.find_header_row(ws)
    return [(ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value)
            for r in range(hr + 1, ws.max_row + 1) if bca.norm(ws.cell(r, 1).value)]


# ---- detection --------------------------------------------------------------

df = sample_df()
_, idc = bca.strip_bom("Response ID"), bca.resolve_id_column(list(df.columns))
check(idc == "Response ID", "resolve_id_column picks the id anchor")
check(bca.resolve_id_column(["ID", "Q1"]) == "ID", "resolve_id_column accepts plain ID")
check(bca.resolve_id_column(["Foo", "Bar"]) == "Foo", "resolve_id_column falls back to first column")

cols, mode = bca.detect_columns(df, idc, columns=["Q1Comment", "xtraNotes"])
check(cols == ["Q1Comment", "xtraNotes"] and mode == "explicit", "explicit columns win, order kept")

cols, mode = bca.detect_columns(df, idc, pattern="comment")
check(cols == ["Q1Comment", "Q9Comment"] and mode == "pattern", "pattern matches *comment* headers")

cols, mode = bca.detect_columns(df, idc)  # default pattern comment|verbatim|feedback
check(cols == ["Q1Comment", "Q9Comment"] and mode == "default-pattern", "default pattern used when nothing given")

cols, mode = bca.detect_columns(df, idc, auto=True)
check(mode == "auto" and "Q1Comment" in cols and "xtraNotes" in cols
      and "Please share your view" in cols, "auto finds free-text incl. oddly-named + question-wording")
check("Q2" not in cols and "Status" not in cols and "Response ID" not in cols,
      "auto excludes categorical, metadata and the id column")

check(bca.comment_pairs(df, idc, "Q1Comment") == [
        (8, "the delivery service is consistently excellent"),
        (32, "they arrive late far too often these days"),
        (55, "helpful and professional staff on every visit"),
        (88, "quick and reliable every time")],
      "comment_pairs keeps only non-blank comments with int ids")


# ---- build + non-destructive incremental update -----------------------------

tmp = Path(tempfile.mkdtemp())
apx = tmp / "apx.xlsx"

s = bca.build_appendix(df, idc, apx, ["Q1Comment", "xtraNotes", "Q9Comment"], "ResponseID")
s["wb"].save(apx)
wb = openpyxl.load_workbook(apx)
check(set(wb.sheetnames) == {"Q1Comment", "xtraNotes", "Q9Comment"}, "one sheet per column")
check(len(sheet_rows(wb["Q1Comment"])) == 4, "Q1Comment got its 4 comments")
check(len(sheet_rows(wb["Q9Comment"])) == 0, "empty column -> empty sheet (header only)")
check(wb["Q1Comment"].cell(bca.find_header_row(wb["Q1Comment"]), 1).value == "ResponseID",
      "new sheet uses the configured id header")

# analyst codes: a theme column + Noteworthy + sentiment on the first Q1Comment row
ws = wb["Q1Comment"]
hr = bca.find_header_row(ws)
ws.cell(hr, 5).value = "Delivery"
ws.cell(hr + 1, 2).value = "p"     # Priority
ws.cell(hr + 1, 4).value = "1"     # sentiment
ws.cell(hr + 1, 5).value = "1"     # theme code
wb.save(apx)

# re-run, same data -> idempotent, coding preserved
wb2 = openpyxl.load_workbook(apx)
s2 = bca.build_appendix(df, idc, apx, ["Q1Comment", "xtraNotes", "Q9Comment"], "ResponseID")
check(s2["added"] == 0, "re-run with same data adds nothing (idempotent)")
ws2 = s2["wb"]["Q1Comment"]
hr2 = bca.find_header_row(ws2)
check(ws2.cell(hr2, 5).value == "Delivery" and ws2.cell(hr2 + 1, 2).value == "p"
      and ws2.cell(hr2 + 1, 4).value == "1" and ws2.cell(hr2 + 1, 5).value == "1",
      "re-run preserves theme column + Priority mark + sentiment")

# new respondent arrives -> appended, coding still intact
df2 = pd.concat([df, pd.DataFrame({"Response ID": [99], "Q1Comment": ["brand new remark"],
                "xtraNotes": [""], "Q9Comment": [""]})], ignore_index=True)
s3 = bca.build_appendix(df2, idc, apx, ["Q1Comment", "xtraNotes", "Q9Comment"], "ResponseID")
ws3 = s3["wb"]["Q1Comment"]
hr3 = bca.find_header_row(ws3)
check(s3["added"] == 1, "a new respondent is appended (added == 1)")
check(ws3.cell(hr3 + 1, 5).value == "1", "the new record does NOT disturb existing coding")
check(any(str(r[0]) == "99" and r[2] == "brand new remark" for r in sheet_rows(ws3)),
      "the new comment lands in the verbatim column")


# ---- SACS-style: id header is 'ID', update finds it -------------------------

apx2 = tmp / "sacs.xlsx"
bca.build_appendix(df, idc, apx2, ["Q1Comment"], "ID")["wb"].save(apx2)
wsS = openpyxl.load_workbook(apx2)["Q1Comment"]
check(wsS.cell(bca.find_header_row(wsS), 1).value == "ID", "id header 'ID' honoured on create")
added, kept = bca.update_sheet(wsS, bca.comment_pairs(df, idc, "Q1Comment"))
check(added == 0 and kept == 4, "update finds the 'ID'-headed sheet and adds nothing")


# ---- changed-comment review (backcheck edits) -------------------------------

tmp2 = Path(tempfile.mkdtemp())
apx3 = tmp2 / "chg.xlsx"
d1 = pd.DataFrame({"Response ID": [1, 2, 3],
                   "QComment": ["original one", "original two", "original three"]})
bca.build_appendix(d1, "Response ID", apx3, ["QComment"], "ResponseID")["wb"].save(apx3)

# analyst codes id 1, and hand-cleans id 2's text in the appendix
wb3 = openpyxl.load_workbook(apx3)
ws3 = wb3["QComment"]
hr3 = bca.find_header_row(ws3)
ws3.cell(hr3 + 1, 2).value = "p"                 # Priority mark on id 1
ws3.cell(hr3 + 2, 3).value = "cleaned two"       # his own edit on id 2
wb3.save(apx3)

# the data now carries a backcheck correction on id 1, and still the raw text on id 2
d2 = pd.DataFrame({"Response ID": [1, 2, 3],
                   "QComment": ["corrected one", "original two", "original three"]})
chg = bca.find_text_changes(d2, "Response ID", ["QComment"], openpyxl.load_workbook(apx3))
found = {(c["id"], c["current"], c["new"]) for c in chg}
check(len(chg) == 2, "find_text_changes finds differences in both directions")
check(("1", "original one", "corrected one") in found, "detects the data-side backcheck correction")
check(("2", "cleaned two", "original two") in found, "detects where the appendix was hand-cleaned")

d3 = pd.DataFrame({"Response ID": [1], "QComment": [""]})
check(bca.find_text_changes(d3, "Response ID", ["QComment"], openpyxl.load_workbook(apx3)) == [],
      "a blank comment in the data is never proposed as a change")

# report -> approve ONLY the backcheck row -> apply
rep = tmp2 / "review.xlsx"
bca.write_change_report(chg, rep)
wbr = openpyxl.load_workbook(rep)
wsr = wbr["Changes"]
check([c.value for c in wsr[1]] == bca.CHANGE_REPORT_HEADERS, "review file carries the expected headers")
for r in range(2, wsr.max_row + 1):
    if str(wsr.cell(r, 2).value).strip() == "1":
        wsr.cell(r, 5).value = "y"
wbr.save(rep)
check(bca.read_change_decisions(rep) == {("QComment", "1")},
      "read_change_decisions returns only the marked row")

wb4 = openpyxl.load_workbook(apx3)
chg4 = bca.find_text_changes(d2, "Response ID", ["QComment"], wb4)
applied = bca.apply_text_changes(wb4, chg4, bca.read_change_decisions(rep))
wb4.save(apx3)
ws5 = openpyxl.load_workbook(apx3)["QComment"]
hr5 = bca.find_header_row(ws5)
check(applied == 1, "apply_text_changes applies only the approved row")
check(ws5.cell(hr5 + 1, 3).value == "corrected one", "approved row took the data text")
check(ws5.cell(hr5 + 2, 3).value == "cleaned two", "unapproved row kept the hand-cleaned text")
check(ws5.cell(hr5 + 1, 2).value == "p", "coding on the updated row is untouched")



# ---- column -> sheet mapping (topic-named appendices) ------------------------
# A hand-built appendix names its sheets for the topic ("Engagement"), not the
# data column ("Q17"). Without a mapping the builder created a SECOND set of
# sheets named after the columns and left the coded ones empty — which looks like
# a successful run and loses nothing, but produces an appendix the report can't
# use and an analyst can't see is wrong.

def config_with_selection(path, rows, header=("QuestionCode", "Include", "CommentSheet")):
    """A minimal crosstab config: title/subtitle block, then the Selection table."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Selection"
    ws.cell(1, 1).value = "Question Selection"
    ws.cell(2, 1).value = "Define which questions appear as stubs and banners."
    for j, h in enumerate(header, start=1):
        ws.cell(3, j).value = h
    ws.cell(4, 1).value = "[REQUIRED] Question code from Survey_Structure."
    for i, row in enumerate(rows, start=5):
        for j, v in enumerate(row, start=1):
            ws.cell(i, j).value = v
    wb.save(path)
    return path


cfg = config_with_selection(tmp / "config.xlsx", [
    ("ResponseID", "N", ""),
    ("Q1Comment", "N", "Delivery talk"),
    ("Q5", "Y", ""),                       # closed question: no CommentSheet, skipped
    ("xtraNotes", "N", "Loose ends"),
])
cols_cfg, map_cfg = bca.load_selection_sheet_map(cfg)
check(cols_cfg == ["Q1Comment", "xtraNotes"],
      "config supplies only the rows carrying a CommentSheet, in order")
check(map_cfg == {"Q1Comment": "Delivery talk", "xtraNotes": "Loose ends"},
      "config supplies the column -> sheet mapping")

# a band-split declaration must REFUSE, not guess a band per comment
cfg_split = config_with_selection(tmp / "config_split.xlsx", [
    ("Q_Rec", "Y", "DetractorComment:Detractor; PromoterComment:Promoter"),
])
try:
    bca.load_selection_sheet_map(cfg_split)
    check(False, "a band-split CommentSheet must refuse")
except bca.SheetMapError as e:
    check("band-split" in str(e) and "Q_Rec" in str(e),
          "the band-split refusal names the offending question")

# a config with no CommentSheet column refuses rather than silently mapping nothing
cfg_nocol = config_with_selection(tmp / "config_nocol.xlsx",
                                  [("Q1Comment", "N")], header=("QuestionCode", "Include"))
try:
    bca.load_selection_sheet_map(cfg_nocol)
    check(False, "a Selection sheet with no CommentSheet column must refuse")
except bca.SheetMapError as e:
    check("CommentSheet" in str(e), "the refusal names the missing column")

check(bca.parse_sheet_map("Q17=Engagement, Q24=Values") == {"Q17": "Engagement", "Q24": "Values"},
      "--sheet-map parses COLUMN=SHEET pairs")
try:
    bca.parse_sheet_map("Q17")
    check(False, "a malformed --sheet-map entry must refuse")
except bca.SheetMapError:
    check(True, "a malformed --sheet-map entry refuses")

check(bca.sheet_name_for("Q17", {"Q17": "Engagement"}) == "Engagement", "mapped column -> its sheet")
check(bca.sheet_name_for("Q99", {"Q17": "Engagement"}) == "Q99", "unmapped column -> its own name")
check(bca.sheet_name_for("Q99", None) == "Q99", "no mapping at all -> unchanged behaviour")


# ---- the mapping preserves coding across a re-run (the whole point) ----------

apx4 = tmp / "topic_named.xlsx"
smap = {"Q1Comment": "Delivery talk", "xtraNotes": "Loose ends"}
cols4 = ["Q1Comment", "xtraNotes"]

bca.build_appendix(df, idc, apx4, cols4, "ID", smap)["wb"].save(apx4)
wb4 = openpyxl.load_workbook(apx4)
check(set(wb4.sheetnames) == {"Delivery talk", "Loose ends"},
      "sheets are created under their MAPPED names")
ws4 = wb4["Delivery talk"]
hr4 = bca.find_header_row(ws4)
check(ws4.cell(hr4, 3).value == "Q1Comment",
      "the verbatim header still records which data column the sheet was built from")

# analyst codes a theme, a tier and a redaction marker
ws4.cell(hr4, 5).value = "Service"
ws4.cell(hr4 + 1, 2).value = "m"        # must-read
ws4.cell(hr4 + 1, 4).value = "3"        # negative
ws4.cell(hr4 + 1, 5).value = "1"
ws4.cell(hr4 + 2, 2).value = "hide"     # redaction: text withheld, still counted
wb4.save(apx4)

# fieldwork grows; re-run must append only, and only into the mapped sheets
df4 = pd.concat([df, pd.DataFrame({"Response ID": [777], "Q1Comment": ["late arrival again"],
                 "xtraNotes": [""], "Q9Comment": [""]})], ignore_index=True)
s4 = bca.build_appendix(df4, idc, apx4, cols4, "ID", smap)
check(s4["added"] == 1, "only the new respondent is appended")
check(s4["created"] == 0 and set(s4["wb"].sheetnames) == {"Delivery talk", "Loose ends"},
      "no duplicate column-named sheets are created on the second run")

ws5 = s4["wb"]["Delivery talk"]
hr5 = bca.find_header_row(ws5)
check(ws5.cell(hr5, 5).value == "Service", "theme COLUMN survives the re-run")
check(ws5.cell(hr5 + 1, 2).value == "m" and ws5.cell(hr5 + 1, 4).value == "3"
      and ws5.cell(hr5 + 1, 5).value == "1", "tier, sentiment and theme coding survive")
check(ws5.cell(hr5 + 2, 2).value == "hide", "a 'hide' redaction marker survives")
check(any(str(r[0]) == "777" and r[2] == "late arrival again" for r in sheet_rows(ws5)),
      "the new comment lands in the mapped sheet")

# the change-review path follows the mapping too
d5 = df4.copy()
d5.loc[d5["Response ID"] == 777, "Q1Comment"] = "late arrival again (corrected)"
s4["wb"].save(apx4)
chg5 = bca.find_text_changes(d5, idc, cols4, openpyxl.load_workbook(apx4), smap)
check(len(chg5) == 1 and chg5[0]["sheet"] == "Delivery talk",
      "find_text_changes reports the MAPPED sheet name")

print("\n" + ("FAILED" if _failed else "OK"), "— %d passed, %d failed" % (_passed, _failed))
raise SystemExit(1 if _failed else 0)
