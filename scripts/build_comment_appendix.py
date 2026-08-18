#!/usr/bin/env python3
"""
build_comment_appendix.py — reusable builder for the Turas qualitative Comment Appendix
=======================================================================================
Turn a survey's open-end columns into a coded-comment workbook the Turas qualitative
tab reads — INCREMENTALLY and NON-DESTRUCTIVELY, so it is safe to re-run as fieldwork
grows without losing any coding you have done.

WHAT IT PRODUCES
----------------
One worksheet per comment column, in the Turas layout. The sheet is named after the
column by default (Q17 -> sheet "Q17"); pass --config (or --sheet-map) when the appendix
names its sheets for the TOPIC instead ("Engagement", "Values"), which hand-built
appendices usually do:

    col A  <id header>         <- join key (matches the survey's respondent id)
    col B  Noteworthy          <- tier code: n=Noteworthy, m=Must-read, p=Priority.
                                  BLANK is the normal state: not flagged, but the
                                  text STILL SHOWS in the report.
                                  "hide" is different — it suppresses the text (the
                                  comment still counts in the numbers). Reserve it
                                  for non-answers ("No.", "Not sure") and gibberish,
                                  NOT for ordinary dull comments: those stay blank.
    col C  <verbatim>          <- the comment text (header = the column name)
    col D  Overall Sentiment   <- coded 1/2/3 (see the legend block above the header)
    col E+ (optional)          <- theme columns you add, coded 1/2/3

A sentiment legend (Total Mentions / 1 Positive / 2 Mixed / 3 Negative) sits above the
header row, matching the hand-built SACS / CCPB appendices.

SAFE TO RE-RUN
--------------
For every sheet that already exists it finds the header row (first cell = the id
anchor), reads the ids already listed, and APPENDS only respondents that are new (have
a comment, not yet present). It NEVER edits or reorders existing rows, so all Noteworthy
marks, Overall Sentiment codes and theme columns are preserved. It writes a timestamped
backup before saving, and refuses (touching nothing) if it cannot find the columns.

CHOOSING THE COMMENT COLUMNS (in priority order)
------------------------------------------------
  --config FILE.xlsx     the crosstab config: every Selection row with a CommentSheet.
                         This also gives each column its SHEET NAME, so one flag covers
                         both questions — and it reads the same declaration the report
                         reads, so the builder and the report cannot disagree.
  --columns "a,b,c"      explicit list (most reliable — use when you know the columns)
  --columns-file FILE    same, one column per line (blank lines / #comments ignored)
  --pattern REGEX        headers matching this regex (default: comment|verbatim|feedback)
  --structure FILE.xlsx  the Survey_Structure's Open_End questions (if the survey tags them)
  --auto                 free-text heuristic (for question-wording headers, e.g. SACS);
                         prints its picks and requires --yes before writing
The resolved columns are always printed with per-column counts and their target sheet,
so you can confirm before anything is written.

WHERE EACH COLUMN'S COMMENTS GO
-------------------------------
  (nothing)              sheet name = column name  (the historical behaviour)
  --config FILE.xlsx     from the Selection sheet's CommentSheet column
  --sheet-map "A=X,B=Y"  explicit; overrides --config for the columns it names
A band-split declaration ('DetractorComment:Detractor; PromoterComment:Promoter') is
REFUSED, not guessed: filing those needs each respondent's score band, and putting a
comment under the wrong band would silently corrupt a hand-coded workbook.

WHEN A COMMENT'S TEXT CHANGES (backcheck edits)
-----------------------------------------------
A normal run never rewrites an existing row, so a comment corrected in the data after the
first build does not flow through. Review those in two steps — a difference can run either
way (the data may hold a correction; the appendix may hold text you cleaned by hand):
  --report-changes [FILE]  write a review list of every differing comment; change nothing
  --apply-changes FILE     rewrite the verbatim ONLY on rows you marked 'y'; coding untouched
Both are surgical: neither appends new respondents.

USAGE
-----
    python3 build_comment_appendix.py --data DATA.xlsx --appendix APX.xlsx --config CONFIG.xlsx
    python3 build_comment_appendix.py --data DATA.xlsx --appendix APX.xlsx --columns "Q1Comment,Q2Comment"
    python3 build_comment_appendix.py --data DATA.xlsx --appendix APX.xlsx --pattern "comment|verbatim|xtra"
    python3 build_comment_appendix.py --data DATA.xlsx --appendix APX.xlsx --auto           # preview
    python3 build_comment_appendix.py --data DATA.xlsx --appendix APX.xlsx --auto --yes      # write
    python3 build_comment_appendix.py ... --dry-run                                          # never write
    python3 build_comment_appendix.py ... --report-changes                                   # review list
    python3 build_comment_appendix.py ... --apply-changes "… changes 20260719_092150.xlsx"

Docs: scripts/README_comment_appendix.md
"""

import argparse
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ---- project preset ---------------------------------------------------------
# A configured copy of this script (one per project, sitting next to its data) may
# fill these in so it runs with no arguments. Any CLI flag overrides them.
DEFAULT_DATA = None          # path to the survey data workbook
DEFAULT_APPENDIX = None      # path to the comment appendix workbook
DEFAULT_COLUMNS = None       # list of comment columns for this project

# ---- layout + detection constants -------------------------------------------

COL_ID, COL_NOTEWORTHY, COL_VERBATIM, COL_SENTIMENT = 1, 2, 3, 4
DEFAULT_ID_HEADER = "ResponseID"
NOTEWORTHY_HEADER = "Noteworthy"
SENTIMENT_HEADER = "Overall Sentiment"

# Header anchor for the respondent-id column — matches the Turas qual reader
# (qual_workbook_reader.R QUAL_ID_PATTERN), so "ID" / "Response ID" / "ResponseID" all work.
ID_PATTERN = re.compile(r"^(response\s*)?id$", re.IGNORECASE)

# Default name pattern for --pattern / a fallback when nothing else is given.
DEFAULT_COMMENT_PATTERN = re.compile(r"comment|verbatim|feedback", re.IGNORECASE)

# Names that are never free-text comments (excluded from the --auto heuristic).
METADATA_DENYLIST = re.compile(
    r"^(response\s*id|id|contact\s*id|.*\bemail\b|.*\bphone\b|.*\burl\b|ip(\s*address)?|"
    r"time\s*started|date\s*submitted|status|longitude|latitude|weight.*|language|country|region)$",
    re.IGNORECASE)

# --auto heuristic thresholds: a free-text column's non-blank cells are long and mostly unique.
AUTO_MIN_MEAN_LEN = 20      # average characters across non-blank cells
AUTO_MIN_UNIQUE_RATIO = 0.6  # distinct / non-blank

SENTIMENT_LEGEND = [
    ("", "", "Total Mentions", ""),
    ("", "1", "Positive skew", ""),
    ("", "2", "Mixed sentiment", ""),
    ("", "3", "Negative skew", ""),
]


# ---- pure helpers (no I/O) --------------------------------------------------

def norm(value):
    """Trimmed string; '' for None/blank."""
    return "" if value is None else str(value).strip()


def strip_bom(name):
    """Drop a leading UTF-8 BOM from a header (a common Alchemer export artifact)."""
    return re.sub("^" + chr(0xFEFF) + "+", "", str(name))


def fmt_id(value):
    """Render a respondent id as the data holds it: integer when whole, else trimmed text."""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        s = value.strip()
        return int(s) if re.fullmatch(r"-?\d+", s) else s
    return value


def resolve_id_column(columns):
    """The respondent-id column: the first header matching the id anchor, else the first column."""
    for c in columns:
        if ID_PATTERN.match(strip_bom(c).strip()):
            return c
    return columns[0]


def looks_like_free_text(series):
    """Heuristic: this column's non-blank values read as free-text comments."""
    vals = [norm(v) for v in series if norm(v) != ""]
    if not vals:
        return False
    mean_len = sum(len(v) for v in vals) / len(vals)
    unique_ratio = len(set(vals)) / len(vals)
    return mean_len >= AUTO_MIN_MEAN_LEN and unique_ratio >= AUTO_MIN_UNIQUE_RATIO


def detect_columns(df, id_col, columns=None, pattern=None, structure_codes=None, auto=False):
    """Resolve the comment columns and the mode used. Returns (ordered_columns, mode).

    Priority: explicit list -> name pattern -> structure Open_End codes -> --auto heuristic.
    Order is preserved and de-duplicated; the id column is never included."""
    headers = [c for c in df.columns if c != id_col]

    if columns:
        wanted = [c.strip() for c in columns if c and c.strip()]
        return list(dict.fromkeys(wanted)), "explicit"
    if pattern is not None:
        rx = pattern if hasattr(pattern, "search") else re.compile(pattern, re.IGNORECASE)
        return [c for c in headers if rx.search(str(c))], "pattern"
    if structure_codes:
        codes = set(structure_codes)
        return [c for c in headers if strip_bom(c).strip() in codes], "structure"
    if auto:
        return [c for c in headers if not METADATA_DENYLIST.match(strip_bom(c).strip())
                and looks_like_free_text(df[c])], "auto"
    # No source given -> the default name pattern (comment|verbatim|feedback).
    return [c for c in headers if DEFAULT_COMMENT_PATTERN.search(str(c))], "default-pattern"


def comment_pairs(df, id_col, column):
    """[(id, comment)] for rows with a non-blank id AND a non-blank comment in `column`."""
    pairs = []
    for rid, comment in zip(df[id_col], df[column]):
        if pd.isna(rid) or norm(rid) == "":
            continue
        text = "" if pd.isna(comment) else str(comment).strip()
        if text:
            pairs.append((fmt_id(rid), text))
    return pairs


def find_header_row(ws):
    """Row index (1-based) whose first cell is the id anchor, else None."""
    for r in range(1, ws.max_row + 1):
        if ID_PATTERN.match(norm(ws.cell(r, 1).value)):
            return r
    return None


def existing_ids_and_last_row(ws, header_row):
    """(set of existing ids as strings, last row carrying an id in column A)."""
    ids, last = set(), header_row
    for r in range(header_row + 1, ws.max_row + 1):
        v = norm(ws.cell(r, COL_ID).value)
        if v:
            ids.add(v)
            last = r
    return ids, last


# ---- column -> sheet mapping ------------------------------------------------
# A sheet is normally named after its data column (Q17 -> sheet "Q17"). Hand-built
# appendices are often named for the TOPIC instead ("Engagement", "Values"), and
# that naming is already declared, per question, in the crosstab config's Selection
# sheet CommentSheet column — which is what the report itself reads to find a
# question's comments. Deriving the mapping from there means the builder and the
# report can never disagree about which sheet belongs to which question.

class SheetMapError(Exception):
    """A mapping that cannot be honoured safely (refuse; never guess)."""


def _selection_header_row(rows, anchor="QuestionCode", limit=10):
    """Index of the Selection header row (it sits under a title/subtitle block)."""
    for i, row in enumerate(rows[:limit]):
        if any(norm(c) == anchor for c in row):
            return i
    return None


def load_selection_sheet_map(config_path):
    """{data column -> appendix sheet} from a crosstab config's Selection sheet.

    Returns (columns_in_display_order, mapping). Only rows with a non-blank
    CommentSheet are included, so the config alone resolves BOTH which columns are
    open-ends and what their sheets are called.

    Refuses on the band-split form ('Detractor:D; Passive:P'), where one question's
    comments are meant to be split across several sheets by a score band. Appending
    that correctly needs the band per respondent; guessing would file comments under
    the wrong band and silently corrupt a hand-coded workbook.
    """
    wb = openpyxl.load_workbook(config_path, data_only=True, read_only=True)
    if "Selection" not in wb.sheetnames:
        raise SheetMapError("no 'Selection' sheet in %s" % Path(config_path).name)
    rows = [list(r) for r in wb["Selection"].iter_rows(values_only=True)]
    hr = _selection_header_row(rows)
    if hr is None:
        raise SheetMapError("no 'QuestionCode' header in the Selection sheet of %s"
                            % Path(config_path).name)
    header = [norm(c) for c in rows[hr]]
    if "CommentSheet" not in header:
        raise SheetMapError(
            "the Selection sheet has no 'CommentSheet' column, so it cannot say which "
            "appendix sheet each open-end belongs to. Add the column, or pass --sheet-map.")
    i_code, i_sheet = header.index("QuestionCode"), header.index("CommentSheet")

    columns, mapping, split = [], {}, []
    for row in rows[hr + 1:]:
        if len(row) <= max(i_code, i_sheet):
            row = list(row) + [None] * (max(i_code, i_sheet) + 1 - len(row))
        code, sheet = norm(row[i_code]), norm(row[i_sheet])
        if not code or not sheet or code.startswith("["):
            continue
        if ";" in sheet or ":" in sheet:
            split.append("%s -> %s" % (code, sheet))
            continue
        if code not in mapping:
            columns.append(code)
            mapping[code] = sheet
    if split:
        raise SheetMapError(
            "band-split open-ends are not supported by this builder:\n    "
            + "\n    ".join(split)
            + "\n  Those questions spread one open-end across several sheets by score band, "
              "which needs each respondent's band to file correctly. Build those sheets by "
              "hand, or drop the band syntax and use one sheet per question.")
    if not mapping:
        raise SheetMapError("no Selection row has a CommentSheet value in %s"
                            % Path(config_path).name)
    return columns, mapping


def parse_sheet_map(text):
    """{col -> sheet} from 'Q17=Engagement,Q24=Values' (explicit alternative to --config)."""
    mapping = {}
    for part in str(text).split(","):
        part = part.strip()
        if not part:
            continue
        if "=" not in part:
            raise SheetMapError("--sheet-map entry %r is not COLUMN=SHEET" % part)
        col, sheet = part.split("=", 1)
        col, sheet = col.strip(), sheet.strip()
        if not col or not sheet:
            raise SheetMapError("--sheet-map entry %r is not COLUMN=SHEET" % part)
        mapping[col] = sheet
    if not mapping:
        raise SheetMapError("--sheet-map is empty")
    return mapping


def sheet_name_for(col, sheet_map):
    """The appendix sheet a data column writes to (itself when unmapped)."""
    return (sheet_map or {}).get(col, col)



# ---- workbook I/O -----------------------------------------------------------

def load_data(path):
    """(df with BOM-stripped headers, id column name)."""
    # pandas' DEFAULT null list contains the strings "None", "NA", "N/A",
    # "NULL" and "nan" - all of which are things a respondent genuinely types
    # in an open end to mean "nothing". Reading with the default list on drops
    # those verbatims silently: the respondent is counted as having said
    # nothing, and their row never reaches the appendix. Only a truly empty
    # cell is missing.
    df = pd.read_excel(path, keep_default_na=False, na_values=[""])
    df.columns = [strip_bom(c) for c in df.columns]
    return df, resolve_id_column(list(df.columns))


def open_end_codes_from_structure(path):
    """Question codes tagged Variable_Type == 'Open_End' in a Survey_Structure workbook."""
    raw = pd.read_excel(path, sheet_name="Questions", header=None)
    hdr = next((i for i in range(len(raw)) if norm(raw.iloc[i, 0]) == "QuestionCode"), 0)
    q = pd.read_excel(path, sheet_name="Questions", header=hdr)
    q.columns = [norm(c) for c in q.columns]
    is_oe = q["Variable_Type"].astype(str).str.strip().str.lower() == "open_end"
    return [norm(c) for c in q.loc[is_oe, "QuestionCode"].tolist() if norm(c)]


def create_sheet(wb, name, id_header, verbatim_header, records):
    """New sheet: legend block + header row + all comment rows. Returns rows written."""
    ws = wb.create_sheet(title=name[:31])          # Excel caps sheet names at 31 chars
    for i, row in enumerate(SENTIMENT_LEGEND, start=1):
        for c, val in enumerate(row, start=1):
            if val:
                ws.cell(i, c).value = val
    hr = len(SENTIMENT_LEGEND) + 1
    ws.cell(hr, COL_ID).value = id_header
    ws.cell(hr, COL_NOTEWORTHY).value = NOTEWORTHY_HEADER
    ws.cell(hr, COL_VERBATIM).value = verbatim_header
    ws.cell(hr, COL_SENTIMENT).value = SENTIMENT_HEADER
    r = hr + 1
    for rid, comment in records:
        ws.cell(r, COL_ID).value = rid
        ws.cell(r, COL_VERBATIM).value = comment
        r += 1
    return len(records)


def update_sheet(ws, records):
    """Append only ids not already present; leave every existing row untouched.

    Returns (n_added, n_kept). Raises ValueError if the sheet has no id header row."""
    header_row = find_header_row(ws)
    if header_row is None:
        raise ValueError("sheet '%s' has no id header row" % ws.title)
    existing, last = existing_ids_and_last_row(ws, header_row)
    r = last + 1
    added = 0
    for rid, comment in records:
        if norm(rid) in existing:
            continue
        ws.cell(r, COL_ID).value = rid
        ws.cell(r, COL_VERBATIM).value = comment
        existing.add(norm(rid))
        r += 1
        added += 1
    return added, len(existing) - added


# ---- changed-comment review (backcheck edits) -------------------------------
# The builder never rewrites an existing row, so a comment corrected in the data
# after you first built the appendix does NOT flow through. These helpers surface
# those differences for review and then apply only the ones you approve — because a
# difference can run either way: the data may carry a backcheck correction, or your
# appendix may hold text you cleaned by hand. Only a human can tell them apart.

CHANGE_REPORT_HEADERS = ["Question", "ResponseID", "Current text (appendix)",
                         "New text (data)", "Apply? (y)"]


def sheet_rows_by_id(ws):
    """{id -> (row_index, verbatim_text)} for a sheet's data rows; {} if no header row."""
    hr = find_header_row(ws)
    if hr is None:
        return {}
    out = {}
    for r in range(hr + 1, ws.max_row + 1):
        rid = norm(ws.cell(r, COL_ID).value)
        if rid and rid not in out:                       # first occurrence wins
            out[rid] = (r, norm(ws.cell(r, COL_VERBATIM).value))
    return out


def find_text_changes(df, id_col, columns, wb, sheet_map=None):
    """Rows whose comment text in the DATA differs from the text held in the appendix.

    Compares only ids present in BOTH, and only when the data's text is non-blank — the
    builder never proposes blanking a comment you already hold. Returns a list of
    {sheet, id, row, current, new}: the review list."""
    changes = []
    for col in dict.fromkeys(columns):
        sheet = sheet_name_for(col, sheet_map)[:31]
        if sheet not in wb.sheetnames or col not in df.columns:
            continue
        data_map = {}
        for rid, txt in zip(df[id_col], df[col]):
            if pd.isna(rid):
                continue
            data_map[norm(fmt_id(rid))] = "" if pd.isna(txt) else str(txt).strip()
        for rid, (row, current) in sheet_rows_by_id(wb[sheet]).items():
            new = data_map.get(rid)
            if new is None or new == "" or new == current:
                continue
            changes.append({"sheet": sheet, "id": rid, "row": row,
                            "current": current, "new": new})
    return changes


def write_change_report(changes, path):
    """Write the review workbook: one row per difference, with a blank Apply? column."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Changes"
    ws.append(CHANGE_REPORT_HEADERS)
    for c in changes:
        ws.append([c["sheet"], c["id"], c["current"], c["new"], ""])
    for i, w in enumerate([18, 12, 60, 60, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(path)
    return len(changes)


def read_change_decisions(path):
    """{(sheet, id)} for the rows marked in the Apply? column (any non-blank mark)."""
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    out = set()
    for r in rows[1:]:                                   # skip the header row
        if not r or len(r) < 5:
            continue
        sheet, rid, mark = norm(r[0]), norm(r[1]), norm(r[4])
        if sheet and rid and mark:
            out.add((sheet, rid))
    return out


def apply_text_changes(wb, changes, decisions):
    """Rewrite the verbatim cell for approved rows only; every coding column untouched."""
    applied = 0
    for c in changes:
        if (c["sheet"], c["id"]) in decisions:
            wb[c["sheet"]].cell(c["row"], COL_VERBATIM).value = c["new"]
            applied += 1
    return applied


# ---- orchestration ----------------------------------------------------------

def build_appendix(df, id_col, appendix_path, columns, id_header, sheet_map=None):
    """Create/update the appendix in place. Returns a summary dict; caller saves the wb.

    Pure w.r.t. the data (df/columns already resolved); the only side effect is the
    returned openpyxl workbook object, which the caller saves once."""
    appendix = Path(appendix_path)
    if appendix.exists():
        wb = openpyxl.load_workbook(appendix)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    summary = {"created": 0, "updated": 0, "empty": 0, "missing": [], "added": 0, "wb": wb}
    for col in dict.fromkeys(columns):
        if col not in df.columns:
            summary["missing"].append(col)
            continue
        records = comment_pairs(df, id_col, col)
        sheet = sheet_name_for(col, sheet_map)[:31]
        label = col if sheet == col else "%s -> %s" % (col, sheet)
        if sheet in wb.sheetnames:
            added, kept = update_sheet(wb[sheet], records)
            summary["added"] += added
            summary["updated"] += 1
            print("  [update]  %-32s kept %3d, added %3d  (total %d)" % (label, kept, added, kept + added))
        else:
            # New sheet: named for the mapping, but the verbatim header stays the
            # DATA column, so the sheet says which column it was built from.
            n = create_sheet(wb, sheet, id_header, col, records)
            summary["added"] += n
            summary["created"] += 1
            if n == 0:
                summary["empty"] += 1
            print("  [create]  %-32s %s" % (label, "empty sheet" if n == 0 else "%d rows" % n))
    return summary


def reorder_sheets(wb, columns, sheet_map=None):
    """Order sheets to match `columns`; any extras keep their place at the end."""
    wanted = []
    for c in dict.fromkeys(columns):
        name = sheet_name_for(c, sheet_map)[:31]
        if name in wb.sheetnames and name not in wanted:
            wanted.append(name)
    extras = [s for s in wb.sheetnames if s not in wanted]
    wb._sheets = [wb[name] for name in wanted + extras]


def run_report_changes(df, id_col, columns, appendix_path, out_path, sheet_map=None):
    """--report-changes: write the review list of differing comments; write nothing else."""
    appendix = Path(appendix_path)
    if not appendix.exists():
        print("ERROR: appendix not found: %s" % appendix)
        return 2
    changes = find_text_changes(df, id_col, columns, openpyxl.load_workbook(appendix), sheet_map)
    if not changes:
        print("No comment text differs between the data and the appendix — nothing to review.")
        return 0
    if not out_path:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = appendix.with_name("%s changes %s.xlsx" % (appendix.stem, stamp))
    write_change_report(changes, out_path)
    by_sheet = {}
    for c in changes:
        by_sheet[c["sheet"]] = by_sheet.get(c["sheet"], 0) + 1
    print("Comment text differs on %d row(s):" % len(changes))
    for sh in sorted(by_sheet):
        print("   %-24s %d" % (sh, by_sheet[sh]))
    print("\nReview list: %s" % Path(out_path).name)
    print("Mark 'y' in the 'Apply? (y)' column on each row where the DATA version should win;")
    print("leave it blank to keep the appendix text. Then re-run with --apply-changes on that file.")
    return 0


def run_apply_changes(df, id_col, columns, appendix_path, review_path, no_backup, sheet_map=None):
    """--apply-changes: rewrite only the approved verbatims; all coding left untouched."""
    appendix = Path(appendix_path)
    wb = openpyxl.load_workbook(appendix)
    changes = find_text_changes(df, id_col, columns, wb, sheet_map)
    decisions = read_change_decisions(review_path)
    if not decisions:
        print("No rows marked in %s — nothing applied." % Path(review_path).name)
        return 0
    applied = apply_text_changes(wb, changes, decisions)
    if applied == 0:
        print("None of the marked rows still differ — appendix left untouched.")
        return 0
    if not no_backup:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = appendix.with_name("%s (backup %s)%s" % (appendix.stem, stamp, appendix.suffix))
        shutil.copy2(appendix, backup)
        print("Backup:  %s" % backup.name)
    wb.save(appendix)
    print("Applied %d text update(s). Noteworthy / sentiment / theme coding left as-is —"
          " re-check the coding on any comment that changed materially." % applied)
    print("Saved:   %s" % appendix.name)
    return 0


def parse_args(argv=None):
    ap = argparse.ArgumentParser(description="Reusable Turas Comment Appendix builder (incremental, non-destructive).")
    ap.add_argument("--data", default=DEFAULT_DATA, required=DEFAULT_DATA is None,
                    help="survey data workbook (.xlsx)")
    ap.add_argument("--appendix", default=DEFAULT_APPENDIX, required=DEFAULT_APPENDIX is None,
                    help="comment appendix workbook to create/update (.xlsx)")
    ap.add_argument("--report-changes", nargs="?", const="", default=None, metavar="FILE",
                    help="write a review list of comments whose text changed in the data; writes nothing else")
    ap.add_argument("--apply-changes", metavar="FILE",
                    help="apply the rows you marked in a --report-changes file (text only; coding untouched)")
    src = ap.add_mutually_exclusive_group()
    src.add_argument("--columns", help="explicit comma-separated comment columns")
    src.add_argument("--columns-file", help="file with one comment column per line (# and blanks ignored)")
    src.add_argument("--pattern", help="regex; data headers matching it are comment columns")
    src.add_argument("--structure", help="Survey_Structure.xlsx; use its Open_End questions")
    src.add_argument("--auto", action="store_true", help="free-text heuristic (prints picks; needs --yes to write)")
    src.add_argument("--config", help="crosstab config .xlsx; its Selection sheet supplies BOTH the "
                                      "open-end columns and the appendix sheet each one writes to "
                                      "(the CommentSheet column)")
    ap.add_argument("--sheet-map", help="COLUMN=SHEET pairs, comma-separated (e.g. 'Q17=Engagement,Q24=Values') "
                                        "for a topic-named appendix without a config to read")
    ap.add_argument("--id-header", default=DEFAULT_ID_HEADER, help="id header for NEW sheets (default: ResponseID)")
    ap.add_argument("--dry-run", action="store_true", help="report only; never write")
    ap.add_argument("--yes", action="store_true", help="confirm writing when columns came from --auto")
    ap.add_argument("--no-backup", action="store_true", help="skip the timestamped backup")
    return ap.parse_args(argv)


def read_columns_file(path):
    lines = Path(path).read_text(encoding="utf-8").splitlines()
    return [ln.strip() for ln in lines if ln.strip() and not ln.strip().startswith("#")]


def main(argv=None):
    args = parse_args(argv)
    df, id_col = load_data(args.data)
    print("Data:     %s" % args.data)
    print("Appendix: %s" % args.appendix)
    print("ID column in data: '%s'\n" % id_col)

    # The appendix sheet each column writes to. Default (no mapping) = the column
    # name, which is the historical behaviour and what an unnamed appendix gets.
    sheet_map = None
    explicit = None
    if args.config:
        try:
            explicit, sheet_map = load_selection_sheet_map(args.config)
        except SheetMapError as e:
            print("ERROR: %s" % e)
            return 2
        print("Config:   %s" % args.config)
    if args.sheet_map:
        try:
            override = parse_sheet_map(args.sheet_map)
        except SheetMapError as e:
            print("ERROR: %s" % e)
            return 2
        sheet_map = dict(sheet_map or {})
        sheet_map.update(override)          # an explicit --sheet-map wins over the config
        if explicit is None:
            explicit = list(override.keys())

    if args.columns:
        explicit = [c for c in args.columns.split(",")]
    elif args.columns_file:
        explicit = read_columns_file(args.columns_file)
    elif explicit is None and DEFAULT_COLUMNS:
        explicit = list(DEFAULT_COLUMNS)
    structure_codes = open_end_codes_from_structure(args.structure) if args.structure else None
    pattern = args.pattern if args.pattern else None

    columns, mode = detect_columns(df, id_col, columns=explicit, pattern=pattern,
                                   structure_codes=structure_codes, auto=args.auto)
    if not columns:
        print("ERROR: no comment columns resolved (mode: %s). The data may be a raw/unmapped\n"
              "  export, or the pattern/list did not match. Nothing was written." % mode)
        return 2

    print("Comment columns (%d, via %s):" % (len(columns), mode))
    for c in columns:
        n = len(comment_pairs(df, id_col, c)) if c in df.columns else "MISSING"
        target = sheet_name_for(c, sheet_map)
        arrow = "" if target == c else "  -> sheet '%s'" % target
        print("   %-24s %s%s" % (c, ("%s comments" % n) if c in df.columns else "NOT IN DATA", arrow))
    print()

    # Safety: an --auto heuristic guess must be eyeballed before it writes.
    if mode == "auto" and not args.yes and not args.dry_run:
        print("Auto-detected columns above — re-run with --yes to write them (or --columns to be explicit).")
        return 0

    # Changed-comment review modes — both are surgical and never append new respondents.
    if args.apply_changes:
        return run_apply_changes(df, id_col, columns, args.appendix,
                                 args.apply_changes, args.no_backup, sheet_map)
    if args.report_changes is not None:
        return run_report_changes(df, id_col, columns, args.appendix, args.report_changes, sheet_map)

    summary = build_appendix(df, id_col, args.appendix, columns, args.id_header, sheet_map)
    for m in summary["missing"]:
        print("  [skip]    %-20s — NOT FOUND in the data file" % m)
    print("\nSummary: %d created (%d empty), %d updated, %d missing; %d comment rows added." % (
        summary["created"], summary["empty"], summary["updated"], len(summary["missing"]), summary["added"]))

    if args.dry_run:
        print("\n--dry-run: no file written.")
        return 0
    if summary["created"] == 0 and summary["added"] == 0:
        print("\nNo new sheets or rows — appendix left untouched (not re-saved).")
        return 0

    reorder_sheets(summary["wb"], columns, sheet_map)
    appendix = Path(args.appendix)
    if appendix.exists() and not args.no_backup:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = appendix.with_name("%s (backup %s)%s" % (appendix.stem, stamp, appendix.suffix))
        shutil.copy2(appendix, backup)
        print("Backup:  %s" % backup.name)
    summary["wb"].save(appendix)
    print("Saved:   %s" % appendix.name)
    return 0


if __name__ == "__main__":
    sys.exit(main())
