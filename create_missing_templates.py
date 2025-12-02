#!/usr/bin/env python3
"""
Create missing TURAS module templates (regular versions)
Generates Pricing, KeyDriver, and Conjoint configuration templates
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def create_pricing_template(output_file):
    """Create Pricing_Config_Template.xlsx"""
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # ===== Settings Sheet =====
    ws_settings = wb.create_sheet("Settings", 0)

    # Headers
    ws_settings['A1'] = 'Setting'
    ws_settings['B1'] = 'Value'

    # Format headers
    for cell in ['A1', 'B1']:
        ws_settings[cell].font = Font(bold=True, size=11)
        ws_settings[cell].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws_settings[cell].font = Font(bold=True, color='FFFFFF', size=11)

    # Data
    settings_data = [
        ['analysis_method', 'van_westendorp'],
        ['data_file', 'pricing_data.csv'],
        ['output_file', 'pricing_results.xlsx'],
        ['weight_var', ''],
        ['dk_codes', ''],
        ['vw_monotonicity_behavior', 'flag_only'],
        ['gg_monotonicity_behavior', 'smooth'],
        ['segment_vars', ''],
        ['unit_cost', '']
    ]

    for i, row_data in enumerate(settings_data, start=2):
        ws_settings[f'A{i}'] = row_data[0]
        ws_settings[f'B{i}'] = row_data[1]

    # Column widths
    ws_settings.column_dimensions['A'].width = 30
    ws_settings.column_dimensions['B'].width = 40

    # ===== VanWestendorp Sheet =====
    ws_vw = wb.create_sheet("VanWestendorp")

    # Headers
    headers_vw = ['Question', 'ColumnName']
    for i, header in enumerate(headers_vw, start=1):
        cell = ws_vw.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)

    # Data
    vw_data = [
        ['too_cheap', 'vw_too_cheap'],
        ['cheap', 'vw_cheap'],
        ['expensive', 'vw_expensive'],
        ['too_expensive', 'vw_too_expensive']
    ]

    for i, row_data in enumerate(vw_data, start=2):
        ws_vw[f'A{i}'] = row_data[0]
        ws_vw[f'B{i}'] = row_data[1]

    ws_vw.column_dimensions['A'].width = 20
    ws_vw.column_dimensions['B'].width = 25

    # ===== GaborGranger Sheet =====
    ws_gg = wb.create_sheet("GaborGranger")

    # Headers
    headers_gg = ['PricePoint', 'PurchaseIntentColumn']
    for i, header in enumerate(headers_gg, start=1):
        cell = ws_gg.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)

    # Data
    gg_data = [
        [49, 'pi_49'],
        [69, 'pi_69'],
        [89, 'pi_89'],
        [99, 'pi_99']
    ]

    for i, row_data in enumerate(gg_data, start=2):
        ws_gg[f'A{i}'] = row_data[0]
        ws_gg[f'B{i}'] = row_data[1]

    ws_gg.column_dimensions['A'].width = 15
    ws_gg.column_dimensions['B'].width = 25

    wb.save(output_file)
    print(f"Created: {output_file}")


def create_keydriver_template(output_file):
    """Create KeyDriver_Config_Template.xlsx"""
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # ===== Settings Sheet =====
    ws_settings = wb.create_sheet("Settings", 0)

    # Headers
    ws_settings['A1'] = 'Setting'
    ws_settings['B1'] = 'Value'

    # Format headers
    for cell in ['A1', 'B1']:
        ws_settings[cell].font = Font(bold=True, size=11)
        ws_settings[cell].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws_settings[cell].font = Font(bold=True, color='FFFFFF', size=11)

    # Data
    settings_data = [
        ['analysis_name', 'Brand Health Drivers'],
        ['data_file', 'keydriver_data.csv'],
        ['output_file', 'keydriver_results.xlsx'],
        ['min_sample_size', 30]
    ]

    for i, row_data in enumerate(settings_data, start=2):
        ws_settings[f'A{i}'] = row_data[0]
        ws_settings[f'B{i}'] = row_data[1]

    # Column widths
    ws_settings.column_dimensions['A'].width = 20
    ws_settings.column_dimensions['B'].width = 40

    # ===== Variables Sheet =====
    ws_vars = wb.create_sheet("Variables")

    # Headers
    headers = ['VariableName', 'Type', 'Label']
    for i, header in enumerate(headers, start=1):
        cell = ws_vars.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)

    # Data
    vars_data = [
        ['overall_satisfaction', 'Outcome', 'Overall Satisfaction'],
        ['product_quality', 'Driver', 'Product Quality'],
        ['customer_service', 'Driver', 'Customer Service'],
        ['value_for_money', 'Driver', 'Value for Money'],
        ['brand_reputation', 'Driver', 'Brand Reputation']
    ]

    for i, row_data in enumerate(vars_data, start=2):
        ws_vars[f'A{i}'] = row_data[0]
        ws_vars[f'B{i}'] = row_data[1]
        ws_vars[f'C{i}'] = row_data[2]

    ws_vars.column_dimensions['A'].width = 25
    ws_vars.column_dimensions['B'].width = 12
    ws_vars.column_dimensions['C'].width = 30

    wb.save(output_file)
    print(f"Created: {output_file}")


def create_conjoint_template(output_file):
    """Create Conjoint_Config_Template.xlsx"""
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # ===== Settings Sheet =====
    ws_settings = wb.create_sheet("Settings", 0)

    # Headers
    ws_settings['A1'] = 'Setting'
    ws_settings['B1'] = 'Value'

    # Format headers
    for cell in ['A1', 'B1']:
        ws_settings[cell].font = Font(bold=True, size=11)
        ws_settings[cell].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws_settings[cell].font = Font(bold=True, color='FFFFFF', size=11)

    # Data
    settings_data = [
        ['analysis_type', 'choice'],
        ['data_file', 'conjoint_data.csv'],
        ['output_file', 'conjoint_results.xlsx'],
        ['choice_set_column', 'choice_set_id'],
        ['chosen_column', 'chosen'],
        ['respondent_id_column', 'resp_id']
    ]

    for i, row_data in enumerate(settings_data, start=2):
        ws_settings[f'A{i}'] = row_data[0]
        ws_settings[f'B{i}'] = row_data[1]

    # Column widths
    ws_settings.column_dimensions['A'].width = 25
    ws_settings.column_dimensions['B'].width = 30

    # ===== Attributes Sheet =====
    ws_attrs = wb.create_sheet("Attributes")

    # Headers
    headers = ['AttributeName', 'NumLevels', 'LevelNames']
    for i, header in enumerate(headers, start=1):
        cell = ws_attrs.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)

    # Data
    attrs_data = [
        ['Price', 3, '£449, £599, £699'],
        ['Brand', 3, 'Apple, Samsung, Google'],
        ['Storage', 3, '128GB, 256GB, 512GB'],
        ['Battery', 3, '12 hours, 18 hours, 24 hours']
    ]

    for i, row_data in enumerate(attrs_data, start=2):
        ws_attrs[f'A{i}'] = row_data[0]
        ws_attrs[f'B{i}'] = row_data[1]
        ws_attrs[f'C{i}'] = row_data[2]

    ws_attrs.column_dimensions['A'].width = 18
    ws_attrs.column_dimensions['B'].width = 12
    ws_attrs.column_dimensions['C'].width = 40

    wb.save(output_file)
    print(f"Created: {output_file}")


if __name__ == "__main__":
    import os

    # Create templates in /templates directory
    templates_dir = "/home/user/Turas/templates"

    create_pricing_template(os.path.join(templates_dir, "Pricing_Config_Template.xlsx"))
    create_keydriver_template(os.path.join(templates_dir, "KeyDriver_Config_Template.xlsx"))
    create_conjoint_template(os.path.join(templates_dir, "Conjoint_Config_Template.xlsx"))

    print("\nAll regular templates created successfully!")
