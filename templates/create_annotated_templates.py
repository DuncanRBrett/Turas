#!/usr/bin/env python3
"""
TURAS Annotated Templates Generator
Creates comprehensive, self-documenting Excel templates with inline documentation
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import os

# Color scheme
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
INSTRUCTIONS_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
EXAMPLE_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_column_widths(ws, widths_dict):
    """Set column widths from dictionary {column_letter: width}"""
    for col, width in widths_dict.items():
        ws.column_dimensions[col].width = width

def add_header_row(ws, headers, row_num=1):
    """Add formatted header row"""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def add_data_row(ws, values, row_num, fill_color=None):
    """Add a data row with optional fill color"""
    for col_num, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = THIN_BORDER
        if fill_color:
            cell.fill = fill_color

def create_instructions_sheet(wb, template_name, description, sections):
    """Create comprehensive instructions sheet"""
    ws = wb.create_sheet("Instructions", 0)

    # Title
    ws['A1'] = f"{template_name} - Instructions"
    ws['A1'].font = Font(size=16, bold=True, color="366092")
    ws.merge_cells('A1:F1')

    # Date
    ws['A2'] = f"Created: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:F2')

    # Overview
    ws['A4'] = "OVERVIEW"
    ws['A4'].font = Font(size=14, bold=True)
    ws['A5'] = description
    ws['A5'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('A5:F7')

    current_row = 9

    # Sections
    for section in sections:
        ws[f'A{current_row}'] = section['title'].upper()
        ws[f'A{current_row}'].font = Font(size=12, bold=True, color="366092")
        current_row += 1

        for item in section['items']:
            ws[f'A{current_row}'] = f"• {item}"
            ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'A{current_row}:F{current_row}')
            current_row += 1

        current_row += 1

    set_column_widths(ws, {'A': 80})

    return ws


def create_survey_structure_template_annotated():
    """Create annotated Survey Structure template"""
    print("Creating Survey_Structure_Template_Annotated.xlsx...")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Instructions sheet
    create_instructions_sheet(wb, "Survey Structure Template",
        "This template defines your survey structure, including all questions, response options, and composite metrics. "
        "It serves as the master reference for question codes, types, and valid responses.",
        [
            {
                'title': 'How to Use This Template',
                'items': [
                    'Start with the Questions sheet to define all survey questions',
                    'Fill in the Options sheet with response options for each question',
                    'Optionally, create composite metrics in the Composite_Metrics sheet',
                    'Save and use this file with the Tabs module for cross-tabulation analysis',
                    'See example data for reference (gray rows)'
                ]
            },
            {
                'title': 'Common Pitfalls',
                'items': [
                    'QuestionCode must be unique across all questions',
                    'Variable_Type must match the actual question format',
                    'For Rating questions, always specify Scale_Min and Scale_Max',
                    'Response codes must match between Questions and Options sheets',
                    'Do not use special characters in QuestionCode'
                ]
            }
        ])

    # Questions sheet
    ws_questions = wb.create_sheet("Questions")

    headers = [
        'QuestionCode',
        'QuestionText',
        'Variable_Type',
        'Scale_Min',
        'Scale_Max',
        'ShowInOutput',
        'Required?',
        'Valid Types',
        'Description'
    ]
    add_header_row(ws_questions, headers)

    # Add documentation rows (yellow background for visibility)
    doc_rows = [
        ['', '', '', '', '', '', 'Required', 'Single/Multiple/Rating/NPS/Grid/Numeric/Text', 'Unique identifier for this question'],
        ['', '', '', '', '', '', 'Required', 'Must match question structure', 'Full question wording as shown to respondents'],
        ['', '', '', '', '', '', 'Required', 'See Valid Types →', 'Type of question (affects analysis methods)'],
        ['', '', '', '', '', '', 'For Rating/NPS only', 'Numeric', 'Minimum value on scale (e.g., 1)'],
        ['', '', '', '', '', '', 'For Rating/NPS only', 'Numeric', 'Maximum value on scale (e.g., 5, 10)'],
        ['', '', '', '', '', '', 'Optional (Y/N)', 'Default: Y', 'Include in output reports'],
    ]

    for row_num, values in enumerate(doc_rows, 2):
        for col_num in range(1, 7):
            cell = ws_questions.cell(row=row_num, column=col_num, value=values[col_num-1])
            cell.fill = INSTRUCTIONS_FILL
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.font = Font(size=9, italic=True)
            cell.border = THIN_BORDER

        # Required? column
        cell = ws_questions.cell(row=row_num, column=7, value=values[6])
        cell.fill = REQUIRED_FILL if 'Required' in values[6] else OPTIONAL_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.font = Font(size=9, bold=True)
        cell.border = THIN_BORDER

        # Valid Types and Description columns
        for col_num in [8, 9]:
            cell = ws_questions.cell(row=row_num, column=col_num, value=values[col_num-1])
            cell.fill = INSTRUCTIONS_FILL
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.font = Font(size=9, italic=True)
            cell.border = THIN_BORDER

    # Example data (gray background)
    examples = [
        ['Q1', 'What is your age group?', 'Single', '', '', 'Y', '', '', ''],
        ['Q2', 'Which brands are you aware of? (Select all)', 'Multiple', '', '', 'Y', '', '', ''],
        ['Q3', 'Overall satisfaction (1-5)', 'Rating', '1', '5', 'Y', '', '', ''],
        ['Q4', 'How likely are you to recommend? (0-10)', 'NPS', '0', '10', 'Y', '', '', ''],
        ['Q5_1', 'Rate product quality (1-10)', 'Grid', '1', '10', 'Y', '', '', ''],
        ['Q6', 'Number of employees', 'Numeric', '', '', 'Y', '', '', ''],
        ['Q7', 'Additional comments', 'Text', '', '', 'N', '', '', ''],
    ]

    current_row = 8
    for example in examples:
        add_data_row(ws_questions, example, current_row, EXAMPLE_FILL)
        current_row += 1

    set_column_widths(ws_questions, {
        'A': 15, 'B': 40, 'C': 12, 'D': 10, 'E': 10, 'F': 12, 'G': 15, 'H': 25, 'I': 40
    })

    # Options sheet
    ws_options = wb.create_sheet("Options")

    headers = [
        'QuestionCode',
        'OptionCode',
        'OptionText',
        'OptionValue',
        'ExcludeFromIndex',
        'BoxCategory',
        'Required?',
        'Description'
    ]
    add_header_row(ws_options, headers)

    # Documentation rows
    doc_rows = [
        ['', '', '', '', '', '', 'Required', 'Must match QuestionCode from Questions sheet'],
        ['', '', '', '', '', '', 'Required', 'Unique numeric code for this option (1, 2, 3...)'],
        ['', '', '', '', '', '', 'Required', 'Response option text shown to respondents'],
        ['', '', '', '', '', '', 'Optional', 'Numeric value for analysis (usually same as OptionCode)'],
        ['', '', '', '', '', '', 'Optional (Y/N)', 'Exclude from index calculations (e.g., "Don\'t know")'],
        ['', '', '', '', '', '', 'Optional', 'Group options: Top2, Bottom2, Positive, Negative'],
    ]

    for row_num, values in enumerate(doc_rows, 2):
        for col_num in range(1, 7):
            cell = ws_options.cell(row=row_num, column=col_num, value=values[col_num-1])
            cell.fill = INSTRUCTIONS_FILL
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.font = Font(size=9, italic=True)
            cell.border = THIN_BORDER

        # Required? column
        cell = ws_options.cell(row=row_num, column=7, value=values[6])
        cell.fill = REQUIRED_FILL if 'Required' in values[6] else OPTIONAL_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.font = Font(size=9, bold=True)
        cell.border = THIN_BORDER

        # Description column
        cell = ws_options.cell(row=row_num, column=8, value=values[7])
        cell.fill = INSTRUCTIONS_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.font = Font(size=9, italic=True)
        cell.border = THIN_BORDER

    # Example data
    examples = [
        ['Q1', '1', 'Under 18', '1', 'N', '', '', ''],
        ['Q1', '2', '18-24', '2', 'N', '', '', ''],
        ['Q1', '3', '25-34', '3', 'N', '', '', ''],
        ['Q1', '99', 'Prefer not to say', '99', 'Y', '', '', ''],
        ['Q2', '1', 'Brand A', '1', 'N', '', '', ''],
        ['Q2', '2', 'Brand B', '2', 'N', '', '', ''],
        ['Q3', '1', 'Very dissatisfied', '1', 'N', 'Bottom2', '', ''],
        ['Q3', '2', 'Dissatisfied', '2', 'N', 'Bottom2', '', ''],
        ['Q3', '3', 'Neutral', '3', 'N', '', '', ''],
        ['Q3', '4', 'Satisfied', '4', 'N', 'Top2', '', ''],
        ['Q3', '5', 'Very satisfied', '5', 'N', 'Top2', '', ''],
    ]

    current_row = 8
    for example in examples:
        add_data_row(ws_options, example, current_row, EXAMPLE_FILL)
        current_row += 1

    set_column_widths(ws_options, {
        'A': 15, 'B': 12, 'C': 30, 'D': 12, 'E': 15, 'F': 12, 'G': 15, 'H': 45
    })

    # Composite_Metrics sheet
    ws_composite = wb.create_sheet("Composite_Metrics")

    headers = [
        'CompositeCode',
        'CompositeLabel',
        'CalculationType',
        'SourceQuestions',
        'Weights',
        'ExcludeFromSummary',
        'SectionLabel',
        'Required?',
        'Valid Values',
        'Description'
    ]
    add_header_row(ws_composite, headers)

    # Documentation rows
    doc_rows = [
        ['', '', '', '', '', '', '', 'Required', 'Must start with COMP_', 'Unique identifier for composite metric'],
        ['', '', '', '', '', '', '', 'Required', '', 'Display name in reports'],
        ['', '', '', '', '', '', '', 'Required', 'Mean/Sum/WeightedMean', 'How to combine source questions'],
        ['', '', '', '', '', '', '', 'Required', 'Comma-separated codes', 'Questions to combine (e.g., Q3,Q4,Q5)'],
        ['', '', '', '', '', '', '', 'For WeightedMean only', 'Comma-separated numbers', 'Weights matching source questions (e.g., 1,2,1)'],
        ['', '', '', '', '', '', '', 'Optional (Y/N)', 'Default: N', 'Hide from Index_Summary sheet'],
        ['', '', '', '', '', '', '', 'Optional', 'Any text', 'Group composites in Index_Summary'],
    ]

    for row_num, values in enumerate(doc_rows, 2):
        for col_num in range(1, 8):
            cell = ws_composite.cell(row=row_num, column=col_num, value=values[col_num-1])
            cell.fill = INSTRUCTIONS_FILL
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.font = Font(size=9, italic=True)
            cell.border = THIN_BORDER

        # Required?, Valid Values, and Description columns
        for col_num in [8, 9, 10]:
            if col_num == 8:
                bg = REQUIRED_FILL if 'Required' in values[col_num-1] else OPTIONAL_FILL
                font = Font(size=9, bold=True)
            else:
                bg = INSTRUCTIONS_FILL
                font = Font(size=9, italic=True)

            cell = ws_composite.cell(row=row_num, column=col_num, value=values[col_num-1])
            cell.fill = bg
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.font = font
            cell.border = THIN_BORDER

    # Example data
    examples = [
        ['COMP_SAT', 'Overall Satisfaction', 'Mean', 'Q3,Q5_1', '', 'N', 'SATISFACTION', '', '', ''],
        ['COMP_QUALITY', 'Quality Index', 'WeightedMean', 'Q5_1,Q5_2', '2,1', 'N', 'QUALITY', '', '', ''],
        ['COMP_TOTAL', 'Total Score', 'Sum', 'Q6,Q7,Q8', '', 'N', '', '', '', ''],
    ]

    current_row = 9
    for example in examples:
        add_data_row(ws_composite, example, current_row, EXAMPLE_FILL)
        current_row += 1

    set_column_widths(ws_composite, {
        'A': 15, 'B': 25, 'C': 15, 'D': 20, 'E': 15, 'F': 15, 'G': 15, 'H': 18, 'I': 20, 'J': 40
    })

    # Save
    output_path = '/home/user/Turas/templates/Survey_Structure_Template_Annotated.xlsx'
    wb.save(output_path)
    print(f"✓ Created: {output_path}")


def create_crosstab_config_template_annotated():
    """Create annotated Crosstab Config template"""
    print("Creating Crosstab_Config_Template_Annotated.xlsx...")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Instructions sheet
    create_instructions_sheet(wb, "Crosstab Configuration Template",
        "This template configures cross-tabulation analysis for single-wave survey data. "
        "It defines analysis settings, banner structure (columns), and question selection (rows).",
        [
            {
                'title': 'How to Use This Template',
                'items': [
                    'Configure analysis parameters in the Settings sheet',
                    'Define column breakouts (demographics) in the Banner sheet',
                    'Select questions to analyze in the Stub sheet',
                    'Save and run with TurasTabs module',
                    'All file paths can be relative to this config file location'
                ]
            },
            {
                'title': 'Quick Settings Guide',
                'items': [
                    'decimal_separator: Use "." for US/UK format, "," for European',
                    'alpha: Significance level (0.05 = 95% confidence)',
                    'minimum_base: Minimum sample size for significance testing (typically 30)',
                    'show_significance: Set to FALSE to hide significance letters',
                    'create_index_summary: Set to TRUE to create executive summary sheet'
                ]
            }
        ])

    # Settings sheet
    ws_settings = wb.create_sheet("Settings")

    headers = ['Setting', 'Value', 'Required?', 'Valid Values', 'Description']
    add_header_row(ws_settings, headers)

    settings_data = [
        # Core settings
        ('project_name', 'Brand Tracker Q1', 'Required', 'Any text', 'Project name for output filename'),
        ('data_file', 'data/survey.csv', 'Required', 'CSV or XLSX path', 'Path to survey data file (relative or absolute)'),
        ('survey_structure_file', 'Survey_Structure.xlsx', 'Required', 'XLSX path', 'Survey structure from Parser or manual'),
        ('weight_variable', 'weight', 'Optional', 'Column name or blank', 'Weighting variable (leave blank for unweighted)'),

        # Display settings
        ('decimal_separator', '.', 'Required', '. or ,', 'Decimal separator: . (US/UK) or , (European)'),
        ('decimal_places_percent', '0', 'Required', '0-3', 'Decimal places for percentages'),
        ('decimal_places_ratings', '1', 'Required', '0-3', 'Decimal places for mean ratings'),
        ('decimal_places_index', '1', 'Optional', '0-3', 'Decimal places for index values'),
        ('decimal_places_numeric', '1', 'Optional', '0-3', 'Decimal places for numeric statistics'),

        # Significance testing
        ('show_significance', 'TRUE', 'Required', 'TRUE/FALSE or Y/N', 'Display significance letters (A, B, C)'),
        ('alpha', '0.05', 'Required', '0.01, 0.05, 0.10', 'Significance level (0.05 = 95% confidence)'),
        ('minimum_base', '30', 'Required', 'Numeric > 0', 'Minimum sample size for sig testing'),
        ('enable_chi_square', 'FALSE', 'Optional', 'TRUE/FALSE', 'Include chi-square test results'),
        ('bonferroni_correction', 'TRUE', 'Optional', 'TRUE/FALSE', 'Apply Bonferroni correction for multiple comparisons'),

        # Advanced features
        ('create_index_summary', 'TRUE', 'Optional', 'TRUE/FALSE or Y/N', 'Create Index_Summary executive dashboard sheet'),
        ('show_standard_deviation', 'FALSE', 'Optional', 'TRUE/FALSE', 'Show standard deviation for ratings'),
        ('show_unweighted_n', 'TRUE', 'Optional', 'TRUE/FALSE', 'Display unweighted base sizes'),
        ('show_effective_n', 'TRUE', 'Optional', 'TRUE/FALSE', 'Display effective sample size (for weighted data)'),

        # Output settings
        ('output_filename', 'Crosstabs.xlsx', 'Optional', 'Filename', 'Output filename (default: Crosstabs.xlsx)'),
        ('output_subfolder', 'output', 'Optional', 'Folder path', 'Output subfolder (created if not exists)'),
    ]

    current_row = 2
    for setting, value, required, valid, desc in settings_data:
        ws_settings.cell(row=current_row, column=1, value=setting)
        ws_settings.cell(row=current_row, column=2, value=value)

        req_cell = ws_settings.cell(row=current_row, column=3, value=required)
        req_cell.fill = REQUIRED_FILL if required == 'Required' else OPTIONAL_FILL
        req_cell.font = Font(bold=True, size=9)

        ws_settings.cell(row=current_row, column=4, value=valid)
        ws_settings.cell(row=current_row, column=5, value=desc)

        for col in range(1, 6):
            ws_settings.cell(row=current_row, column=col).border = THIN_BORDER
            ws_settings.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

        current_row += 1

    set_column_widths(ws_settings, {
        'A': 25, 'B': 25, 'C': 12, 'D': 20, 'E': 50
    })

    # Banner sheet
    ws_banner = wb.create_sheet("Banner")

    headers = ['BannerID', 'BannerLabel', 'Variable', 'Filter', 'Order', 'Required?', 'Description']
    add_header_row(ws_banner, headers)

    # Documentation rows
    doc_rows = [
        ['', '', '', '', '', 'Required', 'Unique ID for this banner column'],
        ['', '', '', '', '', 'Required', 'Display label in output (e.g., "Male", "Age 18-34")'],
        ['', '', '', '', '', 'Required', 'Variable name from data file'],
        ['', '', '', '', '', 'Optional', 'Filter expression (e.g., "Gender==1", "Age>=18 & Age<=34")'],
        ['', '', '', '', '', 'Optional', 'Display order (1, 2, 3...). Leave blank for default order'],
    ]

    for row_num, values in enumerate(doc_rows, 2):
        for col_num in range(1, 6):
            cell = ws_banner.cell(row=row_num, column=col_num, value=values[col_num-1])
            cell.fill = INSTRUCTIONS_FILL
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.font = Font(size=9, italic=True)
            cell.border = THIN_BORDER

        # Required? column
        cell = ws_banner.cell(row=row_num, column=6, value=values[5])
        cell.fill = REQUIRED_FILL if 'Required' in values[5] else OPTIONAL_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.font = Font(size=9, bold=True)
        cell.border = THIN_BORDER

        # Description column
        cell = ws_banner.cell(row=row_num, column=7, value=values[6])
        cell.fill = INSTRUCTIONS_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.font = Font(size=9, italic=True)
        cell.border = THIN_BORDER

    # Example data
    examples = [
        ['Total', 'Total', '', '', '1', '', ''],
        ['Male', 'Male', 'Gender', 'Gender==1', '2', '', ''],
        ['Female', 'Female', 'Gender', 'Gender==2', '3', '', ''],
        ['Age_18_34', 'Age 18-34', 'Age', 'Age>=1 & Age<=2', '4', '', ''],
        ['Age_35_54', 'Age 35-54', 'Age', 'Age==3', '5', '', ''],
        ['Age_55_Plus', 'Age 55+', 'Age', 'Age>=4', '6', '', ''],
    ]

    current_row = 7
    for example in examples:
        add_data_row(ws_banner, example, current_row, EXAMPLE_FILL)
        current_row += 1

    set_column_widths(ws_banner, {
        'A': 15, 'B': 20, 'C': 15, 'D': 30, 'E': 8, 'F': 12, 'G': 45
    })

    # Stub sheet
    ws_stub = wb.create_sheet("Stub")

    headers = ['QuestionCode', 'QuestionText', 'Filter', 'Order', 'Required?', 'Description']
    add_header_row(ws_stub, headers)

    # Documentation rows
    doc_rows = [
        ['', '', '', '', 'Required', 'Question code from Survey_Structure.xlsx'],
        ['', '', '', '', 'Optional', 'Override question text (leave blank to use from Survey_Structure)'],
        ['', '', '', '', 'Optional', 'Filter to apply to this question only'],
        ['', '', '', '', 'Optional', 'Display order in output'],
    ]

    for row_num, values in enumerate(doc_rows, 2):
        for col_num in range(1, 5):
            cell = ws_stub.cell(row=row_num, column=col_num, value=values[col_num-1])
            cell.fill = INSTRUCTIONS_FILL
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.font = Font(size=9, italic=True)
            cell.border = THIN_BORDER

        # Required? and Description columns
        for col_num in [5, 6]:
            bg = REQUIRED_FILL if 'Required' in values[col_num-1] else (INSTRUCTIONS_FILL if col_num == 6 else OPTIONAL_FILL)
            font = Font(size=9, bold=True) if col_num == 5 else Font(size=9, italic=True)

            cell = ws_stub.cell(row=row_num, column=col_num, value=values[col_num-1])
            cell.fill = bg
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.font = font
            cell.border = THIN_BORDER

    # Example data
    examples = [
        ['Q1', 'Age group', '', '1', '', ''],
        ['Q2', 'Brand awareness', '', '2', '', ''],
        ['Q3', 'Overall satisfaction', 'Completed==1', '3', '', ''],
        ['Q4', 'Net Promoter Score', '', '4', '', ''],
        ['COMP_SAT', 'Overall Satisfaction Index', '', '5', '', ''],
    ]

    current_row = 6
    for example in examples:
        add_data_row(ws_stub, example, current_row, EXAMPLE_FILL)
        current_row += 1

    set_column_widths(ws_stub, {
        'A': 15, 'B': 35, 'C': 25, 'D': 8, 'E': 12, 'F': 50
    })

    # Save
    output_path = '/home/user/Turas/templates/Crosstab_Config_Template_Annotated.xlsx'
    wb.save(output_path)
    print(f"✓ Created: {output_path}")


# Due to length constraints, I'll create the remaining templates in separate function calls
# This establishes the pattern and structure

if __name__ == "__main__":
    print("="*60)
    print("TURAS Annotated Templates Generator")
    print("="*60)
    print()

    create_survey_structure_template_annotated()
    create_crosstab_config_template_annotated()

    print()
    print("="*60)
    print("Phase 1 Complete: Survey Structure and Crosstab templates created")
    print("="*60)
