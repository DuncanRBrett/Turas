#!/usr/bin/env python3
"""
TURAS Confidence and Segment Templates Generator
Creates annotated templates for advanced analysis modules
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Color scheme
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
INSTRUCTIONS_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
EXAMPLE_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_column_widths(ws, widths_dict):
    for col, width in widths_dict.items():
        ws.column_dimensions[col].width = width

def add_header_row(ws, headers, row_num=1):
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def add_data_row(ws, values, row_num, fill_color=None):
    for col_num, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = THIN_BORDER
        if fill_color:
            cell.fill = fill_color

def create_instructions_sheet(wb, template_name, description, sections):
    ws = wb.create_sheet("Instructions", 0)

    ws['A1'] = f"{template_name} - Instructions"
    ws['A1'].font = Font(size=16, bold=True, color="366092")
    ws.merge_cells('A1:F1')

    ws['A2'] = f"Created: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:F2')

    ws['A4'] = "OVERVIEW"
    ws['A4'].font = Font(size=14, bold=True)
    ws['A5'] = description
    ws['A5'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('A5:F7')

    current_row = 9

    for section in sections:
        ws[f'A{current_row}'] = section['title'].upper()
        ws[f'A{current_row}'].font = Font(size=12, bold=True, color="366092")
        current_row += 1

        for item in section['items']:
            ws[f'A{current_row}'] = f"• {item}"
            ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'A{current_row}:F{current_row}')
            current_row += 1

        current_row += 1

    set_column_widths(ws, {'A': 80})
    return ws


def create_confidence_config_template_annotated():
    """Create annotated Confidence Config template"""
    print("Creating Confidence_Config_Template_Annotated.xlsx...")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Instructions sheet
    create_instructions_sheet(wb, "Confidence Analysis Configuration Template",
        "This template configures confidence interval analysis for survey data. It supports multiple methods: "
        "Margin of Error (MOE), Wilson Score, Bootstrap, and Bayesian Credible Intervals. "
        "Configure study-wide settings and question-specific analysis parameters.",
        [
            {
                'title': 'Analysis Methods Overview',
                'items': [
                    'MOE (Margin of Error): Classic ±% confidence intervals using normal approximation',
                    'Wilson Score: More accurate for proportions, especially with small samples or extreme values',
                    'Bootstrap: Resampling method (1000-10000 iterations) for robust confidence intervals',
                    'Bayesian Credible Intervals: Incorporates prior beliefs (requires prior specification)'
                ]
            },
            {
                'title': 'When to Use Each Method',
                'items': [
                    'Use MOE for quick estimates and standard reporting',
                    'Use Wilson when sample is small (<30) or proportions are near 0% or 100%',
                    'Use Bootstrap for complex metrics or non-normal distributions',
                    'Use Bayesian when you have prior information from past studies'
                ]
            },
            {
                'title': 'Common Settings',
                'items': [
                    'Confidence_Level: Usually 0.95 (95%), can be 0.90 or 0.99',
                    'Bootstrap_Iterations: 5000 is good balance of speed/accuracy',
                    'Multiple_Comparison_Adjustment: Use when analyzing many questions simultaneously',
                    'DEFF (Design Effect): Set if using complex sampling (usually 1.0-2.0)'
                ]
            }
        ])

    # Study_Settings sheet
    ws_settings = wb.create_sheet("Study_Settings")

    headers = ['Setting', 'Value', 'Required?', 'Valid Values', 'Description']
    add_header_row(ws_settings, headers)

    settings_data = [
        # Core settings
        ('Data_File', 'data/survey.csv', 'Required', 'CSV or XLSX path', 'Path to survey data file'),
        ('Output_File', 'output/confidence_analysis.xlsx', 'Required', 'XLSX path', 'Output file path (will be created)'),

        # Analysis settings
        ('Calculate_Effective_N', 'Y', 'Required', 'Y/N', 'Calculate effective sample size (for weighted data)'),
        ('Confidence_Level', '0.95', 'Required', '0.90, 0.95, 0.99', 'Confidence level (0.95 = 95% confidence intervals)'),
        ('DEFF', '1.0', 'Optional', 'Numeric ≥ 1.0', 'Design effect for complex sampling (1.0 = simple random sample)'),

        # Bootstrap settings
        ('Bootstrap_Iterations', '5000', 'Required', '1000-10000', 'Number of bootstrap resamples (more = more accurate but slower)'),
        ('Random_Seed', '123', 'Optional', 'Any integer', 'Random seed for reproducible bootstrap results'),

        # Multiple comparison adjustment
        ('Multiple_Comparison_Adjustment', 'N', 'Required', 'Y/N', 'Adjust for multiple comparisons (recommended for >10 questions)'),
        ('Multiple_Comparison_Method', 'Bonferroni', 'If Adjustment=Y', 'Bonferroni/Holm/FDR', 'Adjustment method: Bonferroni (conservative), Holm (less conservative), FDR (least conservative)'),

        # Display settings
        ('Decimal_Separator', '.', 'Required', '. or ,', 'Decimal separator: . (US/UK) or , (European)'),
    ]

    current_row = 2
    for setting, value, required, valid, desc in settings_data:
        ws_settings.cell(row=current_row, column=1, value=setting)
        ws_settings.cell(row=current_row, column=2, value=value)

        req_cell = ws_settings.cell(row=current_row, column=3, value=required)
        req_cell.fill = REQUIRED_FILL if required == 'Required' else OPTIONAL_FILL
        req_cell.font = Font(bold=True, size=9)

        ws_settings.cell(row=current_row, column=4, value=valid)
        ws_settings.cell(row=current_row, column=5, value=desc)

        for col in range(1, 6):
            ws_settings.cell(row=current_row, column=col).border = THIN_BORDER
            ws_settings.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

        current_row += 1

    set_column_widths(ws_settings, {
        'A': 30, 'B': 25, 'C': 15, 'D': 20, 'E': 55
    })

    # Question_Analysis sheet
    ws_questions = wb.create_sheet("Question_Analysis")

    headers = [
        'Question_ID',
        'Statistic_Type',
        'Categories',
        'Run_MOE',
        'Run_Bootstrap',
        'Run_Credible',
        'Use_Wilson',
        'Prior_Mean',
        'Prior_SD',
        'Promoter_Codes',
        'Detractor_Codes',
        'Required?',
        'Valid Values',
        'Description'
    ]
    add_header_row(ws_questions, headers)

    # Documentation rows
    doc_rows = [
        ['', '', '', '', '', '', '', '', '', '', '', 'Required', 'Any code', 'Question code from data file'],
        ['', '', '', '', '', '', '', '', '', '', '', 'Required', 'proportion/mean/nps', 'Type of statistic to analyze'],
        ['', '', '', '', '', '', '', '', '', '', '', 'For proportion only', 'Comma-separated', 'Response codes to include (e.g., 1,2 for "Yes,No")'],
        ['', '', '', '', '', '', '', '', '', '', '', 'Optional', 'Y/N', 'Calculate classic Margin of Error (±%)'],
        ['', '', '', '', '', '', '', '', '', '', '', 'Optional', 'Y/N', 'Calculate bootstrap confidence interval'],
        ['', '', '', '', '', '', '', '', '', '', '', 'Optional', 'Y/N', 'Calculate Bayesian credible interval'],
        ['', '', '', '', '', '', '', '', '', '', '', 'For proportion only', 'Y/N', 'Use Wilson score interval (more accurate for small samples)'],
        ['', '', '', '', '', '', '', '', '', '', '', 'For Bayesian only', 'Numeric', 'Prior mean estimate (from past studies)'],
        ['', '', '', '', '', '', '', '', '', '', '', 'For Bayesian mean/nps', 'Numeric > 0', 'Prior standard deviation (uncertainty in prior)'],
        ['', '', '', '', '', '', '', '', '', '', '', 'For NPS only', 'Comma-separated', 'Codes for promoters (typically 9,10)'],
        ['', '', '', '', '', '', '', '', '', '', '', 'For NPS only', 'Comma-separated', 'Codes for detractors (typically 0,1,2,3,4,5,6)'],
    ]

    for row_num, values in enumerate(doc_rows, 2):
        for col_num in range(1, 12):
            cell = ws_questions.cell(row=row_num, column=col_num, value=values[col_num-1])
            cell.fill = INSTRUCTIONS_FILL
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.font = Font(size=9, italic=True)
            cell.border = THIN_BORDER

        # Required?, Valid Values, and Description columns
        for col_num in [12, 13, 14]:
            if col_num == 12:
                bg = REQUIRED_FILL if 'Required' in values[col_num-1] else OPTIONAL_FILL
                font = Font(size=9, bold=True)
            else:
                bg = INSTRUCTIONS_FILL
                font = Font(size=9, italic=True)

            cell = ws_questions.cell(row=row_num, column=col_num, value=values[col_num-1])
            cell.fill = bg
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.font = font
            cell.border = THIN_BORDER

    # Example data
    examples = [
        ['Q1', 'proportion', '1', 'Y', 'Y', 'N', 'Y', '', '', '', '', '', '', ''],
        ['Q2', 'proportion', '4,5', 'Y', 'Y', 'Y', 'N', '0.65', '', '', '', '', '', ''],
        ['Q3', 'mean', '', 'Y', 'Y', 'Y', '', '7.5', '1.2', '', '', '', '', ''],
        ['Q4', 'nps', '', 'Y', 'Y', 'N', '', '', '', '9,10', '0,1,2,3,4,5,6', '', '', ''],
        ['Q5', 'mean', '', 'Y', 'N', 'N', '', '', '', '', '', '', '', ''],
    ]

    current_row = 13
    for example in examples:
        add_data_row(ws_questions, example, current_row, EXAMPLE_FILL)
        current_row += 1

    set_column_widths(ws_questions, {
        'A': 12, 'B': 12, 'C': 12, 'D': 10, 'E': 12, 'F': 12, 'G': 12,
        'H': 10, 'I': 10, 'J': 15, 'K': 20, 'L': 15, 'M': 18, 'N': 40
    })

    # Save
    output_path = '/home/user/Turas/templates/Confidence_Config_Template_Annotated.xlsx'
    wb.save(output_path)
    print(f"✓ Created: {output_path}")


def create_segment_config_template_annotated():
    """Create annotated Segment Config template"""
    print("Creating Segment_Config_Template_Annotated.xlsx...")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Instructions sheet
    create_instructions_sheet(wb, "Segmentation Analysis Configuration Template",
        "This template configures k-means clustering segmentation analysis. It supports two modes: "
        "Exploration (test multiple k values) and Final (run with fixed k). "
        "Configure clustering variables, data handling, outlier detection, and output preferences.",
        [
            {
                'title': 'Analysis Modes',
                'items': [
                    'Exploration Mode: Set k_fixed to blank. System tests k_min to k_max and recommends best k',
                    'Final Run Mode: Set k_fixed to specific number (e.g., 4). System creates final segments',
                    'Recommendation: Start with exploration, review metrics, then run final with chosen k'
                ]
            },
            {
                'title': 'Key Configuration Decisions',
                'items': [
                    'clustering_vars: Choose 5-15 discriminating variables (too few = poor separation, too many = noise)',
                    'standardize: Usually TRUE (scales variables to same range)',
                    'missing_data: listwise_deletion (default), mean_imputation, or median_imputation',
                    'outlier_detection: Enable to identify unusual respondents (zscore or mahalanobis methods)',
                    'k_min/k_max: Typically 3-6 for consumer segmentation, 2-4 for B2B'
                ]
            },
            {
                'title': 'Output Files',
                'items': [
                    'Creates dated folder (YYYYMMDD) in output directory',
                    'segment_assignments.csv: Respondent-level segment membership',
                    'segment_profiles.xlsx: Detailed segment characteristics',
                    'segment_validation.xlsx: Quality metrics (silhouette, elbow, gap)',
                    'segment_model.rds: Saved model (for scoring new data)'
                ]
            }
        ])

    # Config sheet
    ws = wb.create_sheet("Config")

    headers = ['Parameter', 'Value', 'Required?', 'Valid Values', 'Description']
    add_header_row(ws, headers)

    config_data = [
        # Data source
        ('data_file', 'data/survey.xlsx', 'Required', 'CSV or XLSX path', 'Path to survey data file'),
        ('data_sheet', 'Data', 'Required', 'Sheet name', 'Sheet name in Excel (or leave as "Data" for CSV)'),
        ('id_variable', 'respondent_id', 'Required', 'Column name', 'Unique identifier column for respondents'),

        # Variables
        ('clustering_vars', 'Q1,Q2,Q3,Q4,Q5,Q6', 'Required', 'Comma-separated codes', 'Variables to use for clustering (5-15 recommended)'),
        ('profile_vars', 'Gender,Age,Region', 'Optional', 'Comma-separated codes or blank', 'Variables for profiling (leave blank to use all non-clustering vars)'),

        # Model configuration
        ('method', 'kmeans', 'Required', 'kmeans', 'Clustering method (currently only kmeans supported)'),
        ('k_fixed', '', 'Optional', 'Integer 2-10 or blank', 'Fixed k for final run (blank = exploration mode)'),
        ('k_min', '3', 'For exploration', '2-10', 'Minimum k to test (exploration mode)'),
        ('k_max', '6', 'For exploration', '2-15', 'Maximum k to test (exploration mode)'),
        ('nstart', '50', 'Required', '1-200', 'Number of random starts (50-100 recommended for stability)'),
        ('seed', '123', 'Required', 'Any integer', 'Random seed for reproducibility'),

        # Data handling
        ('missing_data', 'listwise_deletion', 'Required', 'listwise_deletion/mean_imputation/median_imputation/refuse', 'How to handle missing values'),
        ('missing_threshold', '15', 'Optional', '0-100', 'Maximum % missing per variable (variables exceeding this are excluded)'),
        ('standardize', 'TRUE', 'Required', 'TRUE/FALSE', 'Standardize variables to mean=0, sd=1 (usually TRUE)'),
        ('min_segment_size_pct', '10', 'Optional', '0-50', 'Minimum segment size as % of sample (segments smaller flagged as warning)'),

        # Outlier detection
        ('outlier_detection', 'FALSE', 'Optional', 'TRUE/FALSE', 'Enable outlier detection (identifies unusual respondents)'),
        ('outlier_method', 'zscore', 'If outlier_detection=TRUE', 'zscore/mahalanobis', 'Outlier detection method: zscore (simple), mahalanobis (multivariate)'),
        ('outlier_threshold', '3.0', 'If outlier_detection=TRUE', '1.0-5.0', 'Threshold: zscore (usually 3.0), mahalanobis (chi-square based)'),
        ('outlier_min_vars', '1', 'If outlier_detection=TRUE', '1-nclustering_vars', 'Min variables flagged for outlier status'),
        ('outlier_handling', 'flag', 'If outlier_detection=TRUE', 'none/flag/remove', 'How to handle outliers: none (ignore), flag (mark), remove (exclude)'),
        ('outlier_alpha', '0.001', 'For mahalanobis', '0.0001-0.1', 'Significance level for mahalanobis test (0.001 = 99.9% confidence)'),

        # Variable selection
        ('variable_selection', 'FALSE', 'Optional', 'TRUE/FALSE', 'Enable automatic variable selection (reduces to max_clustering_vars)'),
        ('variable_selection_method', 'variance_correlation', 'If variable_selection=TRUE', 'variance_correlation/factor_analysis/both', 'Selection method'),
        ('max_clustering_vars', '10', 'If variable_selection=TRUE', '2-20', 'Target number of clustering variables after selection'),
        ('varsel_min_variance', '0.1', 'If variable_selection=TRUE', '0.01-1.0', 'Minimum variance to retain variable'),
        ('varsel_max_correlation', '0.8', 'If variable_selection=TRUE', '0.5-0.95', 'Maximum correlation before removing redundant variable'),

        # Validation
        ('k_selection_metrics', 'silhouette,elbow', 'For exploration', 'silhouette/elbow/gap (comma-separated)', 'Metrics for k selection: silhouette (cluster cohesion), elbow (within-SS), gap (statistical)'),

        # Output
        ('output_folder', 'output/', 'Required', 'Directory path', 'Output directory (created if not exists)'),
        ('output_prefix', 'seg_', 'Optional', 'Any text', 'Prefix for output filenames'),
        ('create_dated_folder', 'TRUE', 'Optional', 'TRUE/FALSE', 'Create subfolder with YYYYMMDD date stamp'),
        ('segment_names', 'auto', 'Optional', 'auto or comma-separated names', 'Segment names (auto = "Segment 1", "Segment 2", etc.)'),
        ('save_model', 'TRUE', 'Optional', 'TRUE/FALSE', 'Save model object (.rds) for scoring new data'),

        # Metadata
        ('project_name', 'Customer Segmentation', 'Optional', 'Any text', 'Project name (for documentation)'),
        ('analyst_name', 'Analyst Name', 'Optional', 'Any text', 'Analyst name (for documentation)'),
        ('description', 'B2C customer segmentation analysis', 'Optional', 'Any text', 'Project description'),

        # Question labels
        ('question_labels_file', 'question_labels.xlsx', 'Optional', 'XLSX path or blank', 'File with variable labels (2 columns: variable, label)'),
    ]

    current_row = 2
    for param, value, required, valid, desc in config_data:
        ws.cell(row=current_row, column=1, value=param)
        ws.cell(row=current_row, column=2, value=value)

        req_cell = ws.cell(row=current_row, column=3, value=required)
        req_cell.fill = REQUIRED_FILL if required == 'Required' else OPTIONAL_FILL
        req_cell.font = Font(bold=True, size=9)

        ws.cell(row=current_row, column=4, value=valid)
        ws.cell(row=current_row, column=5, value=desc)

        for col in range(1, 6):
            ws.cell(row=current_row, column=col).border = THIN_BORDER
            ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

        current_row += 1

    set_column_widths(ws, {
        'A': 28, 'B': 30, 'C': 18, 'D': 35, 'E': 60
    })

    # Save
    output_path = '/home/user/Turas/templates/Segment_Config_Template_Annotated.xlsx'
    wb.save(output_path)
    print(f"✓ Created: {output_path}")


if __name__ == "__main__":
    print("="*60)
    print("TURAS Confidence & Segment Templates Generator")
    print("="*60)
    print()

    create_confidence_config_template_annotated()
    create_segment_config_template_annotated()

    print()
    print("="*60)
    print("Confidence and Segment templates created successfully")
    print("="*60)
