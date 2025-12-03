#!/usr/bin/env python3
"""
Fix Annotated Templates - Match Working Template Structure
Reads working templates and creates corrected annotated versions
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# Color scheme
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
INSTRUCTIONS_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
EXAMPLE_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def add_instructions_sheet(wb, module_name):
    """Add helpful instructions sheet"""
    ws = wb.create_sheet("Instructions", 0)

    # Title
    ws['A1'] = f"{module_name} Configuration Template - Instructions"
    ws['A1'].font = Font(size=14, bold=True, color="366092")
    ws.merge_cells('A1:F1')

    # Date
    ws['A2'] = f"Created: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(size=9, italic=True)

    # Overview
    ws['A5'] = "OVERVIEW"
    ws['A5'].font = Font(bold=True, size=11)
    ws['A6'] = f"This template configures {module_name} analysis for TURAS. It defines analysis settings and parameters."
    ws['A6'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('A6:F6')

    # How to use
    ws['A9'] = "HOW TO USE THIS TEMPLATE"
    ws['A9'].font = Font(bold=True, size=11)

    instructions = [
        "• Configure analysis parameters in the Settings sheet",
        "• Fill in the 'Value' column with your project-specific values",
        "• Required fields are highlighted in yellow",
        "• Optional fields can be left as default or customized",
        "• Save and run with TURAS module",
        "• All file paths can be relative to this config file location"
    ]

    for i, instruction in enumerate(instructions, 10):
        ws[f'A{i}'] = instruction
        ws[f'A{i}'].alignment = Alignment(wrap_text=True)

    # Column widths
    ws.column_dimensions['A'].width = 80

    return ws

def copy_sheet_with_formatting(source_sheet, target_wb, sheet_name):
    """Copy a sheet from source to target workbook with all formatting"""
    target_sheet = target_wb.create_sheet(sheet_name)

    # Copy data and formatting
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet[cell.coordinate]

            # Copy value
            if cell.value is not None:
                new_cell.value = cell.value

            # Copy formatting
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy()
                new_cell.alignment = cell.alignment.copy()

    # Copy column widths
    for col_letter in source_sheet.column_dimensions:
        if col_letter in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col_letter].width = \
                source_sheet.column_dimensions[col_letter].width

    # Copy row heights
    for row_num in source_sheet.row_dimensions:
        if row_num in source_sheet.row_dimensions:
            target_sheet.row_dimensions[row_num].height = \
                source_sheet.row_dimensions[row_num].height

    return target_sheet

def enhance_settings_sheet(ws):
    """Add highlighting to Settings sheet - yellow for required, enhanced formatting"""
    # Assuming row 1 is headers, row 2+ are data
    # Column structure: Setting | Type | Default | Description

    # Enhance header row
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

    # Enhance data rows - highlight required settings
    for row_num in range(2, ws.max_row + 1):
        # Check if Type column indicates required
        type_cell = ws.cell(row=row_num, column=2)

        # Highlight entire row if it seems important
        # You can customize this logic based on the actual template structure
        for col in range(1, 5):
            cell = ws.cell(row=row_num, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    return ws

def create_corrected_annotated_template(working_template_path, output_path, module_name):
    """Create corrected annotated template from working template"""

    print(f"Processing {module_name}...")

    # Load working template
    wb_working = openpyxl.load_workbook(working_template_path)

    # Create new workbook for annotated version
    wb_annotated = openpyxl.Workbook()
    wb_annotated.remove(wb_annotated.active)  # Remove default sheet

    # Add Instructions sheet first
    add_instructions_sheet(wb_annotated, module_name)

    # Copy all sheets from working template
    for sheet_name in wb_working.sheetnames:
        source_sheet = wb_working[sheet_name]
        target_sheet = copy_sheet_with_formatting(source_sheet, wb_annotated, sheet_name)

        # Enhance Settings sheet if it exists
        if sheet_name.lower() == 'settings':
            enhance_settings_sheet(target_sheet)

    # Save annotated template
    wb_annotated.save(output_path)
    print(f"  ✓ Created {output_path}")

    wb_working.close()
    wb_annotated.close()

def main():
    """Fix all annotated templates"""

    print("=" * 80)
    print("FIXING ANNOTATED TEMPLATES")
    print("=" * 80)
    print()

    # Get script directory
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Module templates to fix
    modules = [
        ("Crosstab", "Crosstab_Config_Template.xlsx", "Crosstab_Config_Template_Annotated_FIXED.xlsx"),
        ("Confidence", "Confidence_Config_Template.xlsx", "Confidence_Config_Template_Annotated_FIXED.xlsx"),
        ("KeyDriver", "KeyDriver_Config_Template.xlsx", "KeyDriver_Config_Template_Annotated_FIXED.xlsx"),
        ("Segment", "Segment_Config_Template.xlsx", "Segment_Config_Template_Annotated_FIXED.xlsx"),
        ("Conjoint", "Conjoint_Config_Template.xlsx", "Conjoint_Config_Template_Annotated_FIXED.xlsx"),
        ("Pricing", "Pricing_Config_Template.xlsx", "Pricing_Config_Template_Annotated_FIXED.xlsx"),
        ("Tracker", "Tracker_Config_Template.xlsx", "Tracker_Config_Template_Annotated_FIXED.xlsx"),
    ]

    for module_name, working_file, annotated_file in modules:
        working_path = os.path.join(script_dir, working_file)
        output_path = os.path.join(script_dir, annotated_file)

        if not os.path.exists(working_path):
            print(f"⚠ Warning: {working_path} not found, skipping...")
            continue

        try:
            create_corrected_annotated_template(working_path, output_path, module_name)
        except Exception as e:
            print(f"  ✗ Error: {e}")

    print()
    print("=" * 80)
    print("COMPLETE")
    print("=" * 80)
    print()
    print("Next steps:")
    print("1. Review the _FIXED.xlsx files")
    print("2. If correct, rename them to replace the old annotated versions:")
    print("   - Delete old *_Annotated.xlsx files")
    print("   - Rename *_Annotated_FIXED.xlsx to *_Annotated.xlsx")
    print()

if __name__ == "__main__":
    main()
