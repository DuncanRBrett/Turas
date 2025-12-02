#!/usr/bin/env python3
"""
TURAS Tracker Templates Generator
Creates annotated Tracker configuration templates
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Color scheme
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
INSTRUCTIONS_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
EXAMPLE_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_column_widths(ws, widths_dict):
    """Set column widths from dictionary {column_letter: width}"""
    for col, width in widths_dict.items():
        ws.column_dimensions[col].width = width

def add_header_row(ws, headers, row_num=1):
    """Add formatted header row"""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def add_data_row(ws, values, row_num, fill_color=None):
    """Add a data row with optional fill color"""
    for col_num, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = THIN_BORDER
        if fill_color:
            cell.fill = fill_color

def create_instructions_sheet(wb, template_name, description, sections):
    """Create comprehensive instructions sheet"""
    ws = wb.create_sheet("Instructions", 0)

    # Title
    ws['A1'] = f"{template_name} - Instructions"
    ws['A1'].font = Font(size=16, bold=True, color="366092")
    ws.merge_cells('A1:F1')

    # Date
    ws['A2'] = f"Created: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:F2')

    # Overview
    ws['A4'] = "OVERVIEW"
    ws['A4'].font = Font(size=14, bold=True)
    ws['A5'] = description
    ws['A5'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('A5:F7')

    current_row = 9

    # Sections
    for section in sections:
        ws[f'A{current_row}'] = section['title'].upper()
        ws[f'A{current_row}'].font = Font(size=12, bold=True, color="366092")
        current_row += 1

        for item in section['items']:
            ws[f'A{current_row}'] = f"• {item}"
            ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'A{current_row}:F{current_row}')
            current_row += 1

        current_row += 1

    set_column_widths(ws, {'A': 80})

    return ws


def create_tracker_config_template_annotated():
    """Create annotated Tracker Config template"""
    print("Creating Tracker_Config_Template_Annotated.xlsx...")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Instructions sheet
    create_instructions_sheet(wb, "Tracker Configuration Template",
        "This template configures multi-wave tracking analysis. It defines survey waves, tracked questions, "
        "demographic breakouts, and analysis settings for trend analysis across time periods.",
        [
            {
                'title': 'How to Use This Template',
                'items': [
                    'Define survey waves in the Waves sheet (one row per wave)',
                    'List questions to track in the TrackedQuestions sheet',
                    'Configure banner breakouts (demographics) in the Banner sheet',
                    'Set analysis parameters in the Settings sheet',
                    'Use with Tracker_Question_Mapping to handle question code changes',
                    'Run with TurasTracker module'
                ]
            },
            {
                'title': 'Wave Configuration Tips',
                'items': [
                    'WaveID should be short and consistent (W1, W2, W3 or Q1_2024, Q2_2024)',
                    'DataFile paths can be relative to this config file',
                    'WeightVar must have the same name across all wave data files',
                    'Fieldwork dates are for documentation - not used in calculations'
                ]
            }
        ])

    # Waves sheet
    ws_waves = wb.create_sheet("Waves")

    headers = [
        'WaveID',
        'WaveName',
        'DataFile',
        'FieldworkStart',
        'FieldworkEnd',
        'WeightVar',
        'Required?',
        'Description'
    ]
    add_header_row(ws_waves, headers)

    # Documentation rows
    doc_rows = [
        ['', '', '', '', '', '', 'Required', 'Short unique identifier (e.g., W1, W2, Q1_2024)'],
        ['', '', '', '', '', '', 'Required', 'Descriptive name for reports (e.g., "Wave 1 - Jan 2024")'],
        ['', '', '', '', '', '', 'Required', 'Path to CSV or Excel data file (relative or absolute)'],
        ['', '', '', '', '', '', 'Optional', 'Fieldwork start date (YYYY-MM-DD format)'],
        ['', '', '', '', '', '', 'Optional', 'Fieldwork end date (YYYY-MM-DD format)'],
        ['', '', '', '', '', '', 'Optional', 'Weight variable name (must be consistent across waves)'],
    ]

    for row_num, values in enumerate(doc_rows, 2):
        for col_num in range(1, 7):
            cell = ws_waves.cell(row=row_num, column=col_num, value=values[col_num-1])
            cell.fill = INSTRUCTIONS_FILL
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.font = Font(size=9, italic=True)
            cell.border = THIN_BORDER

        # Required? column
        cell = ws_waves.cell(row=row_num, column=7, value=values[6])
        cell.fill = REQUIRED_FILL if 'Required' in values[6] else OPTIONAL_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.font = Font(size=9, bold=True)
        cell.border = THIN_BORDER

        # Description column
        cell = ws_waves.cell(row=row_num, column=8, value=values[7])
        cell.fill = INSTRUCTIONS_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.font = Font(size=9, italic=True)
        cell.border = THIN_BORDER

    # Example data
    examples = [
        ['W1', 'Wave 1 - Jan 2024', 'data/wave1.csv', '2024-01-15', '2024-01-30', 'weight', '', ''],
        ['W2', 'Wave 2 - Apr 2024', 'data/wave2.csv', '2024-04-15', '2024-04-30', 'weight', '', ''],
        ['W3', 'Wave 3 - Jul 2024', 'data/wave3.csv', '2024-07-15', '2024-07-30', 'weight', '', ''],
        ['W4', 'Wave 4 - Oct 2024', 'data/wave4.csv', '2024-10-15', '2024-10-30', 'weight', '', ''],
    ]

    current_row = 8
    for example in examples:
        add_data_row(ws_waves, example, current_row, EXAMPLE_FILL)
        current_row += 1

    set_column_widths(ws_waves, {
        'A': 10, 'B': 25, 'C': 25, 'D': 15, 'E': 15, 'F': 12, 'G': 12, 'H': 45
    })

    # TrackedQuestions sheet
    ws_tracked = wb.create_sheet("TrackedQuestions")

    headers = ['QuestionCode', 'Required?', 'Description']
    add_header_row(ws_tracked, headers)

    # Documentation rows
    doc_rows = [
        ['', 'Required', 'Standard question code used for tracking (matches question_mapping.xlsx)'],
    ]

    for row_num, values in enumerate(doc_rows, 2):
        cell = ws_tracked.cell(row=row_num, column=1, value=values[0])
        cell.fill = INSTRUCTIONS_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.font = Font(size=9, italic=True)
        cell.border = THIN_BORDER

        cell = ws_tracked.cell(row=row_num, column=2, value=values[1])
        cell.fill = REQUIRED_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.font = Font(size=9, bold=True)
        cell.border = THIN_BORDER

        cell = ws_tracked.cell(row=row_num, column=3, value=values[2])
        cell.fill = INSTRUCTIONS_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.font = Font(size=9, italic=True)
        cell.border = THIN_BORDER

    # Example data
    examples = [
        ['Q_SAT', '', ''],
        ['Q_NPS', '', ''],
        ['Q_VALUE', '', ''],
        ['Q_QUALITY', '', ''],
        ['COMP_OVERALL', '', ''],
    ]

    current_row = 3
    for example in examples:
        add_data_row(ws_tracked, example, current_row, EXAMPLE_FILL)
        current_row += 1

    set_column_widths(ws_tracked, {
        'A': 20, 'B': 12, 'C': 60
    })

    # Banner sheet
    ws_banner = wb.create_sheet("Banner")

    headers = ['BreakVariable', 'BreakLabel', 'Required?', 'Description']
    add_header_row(ws_banner, headers)

    # Documentation rows
    doc_rows = [
        ['', '', 'Required', 'Variable name in data files (must exist in all waves)'],
        ['', '', 'Required', 'Display label for reports'],
    ]

    for row_num, values in enumerate(doc_rows, 2):
        for col_num in range(1, 3):
            cell = ws_banner.cell(row=row_num, column=col_num, value=values[col_num-1])
            cell.fill = INSTRUCTIONS_FILL
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.font = Font(size=9, italic=True)
            cell.border = THIN_BORDER

        cell = ws_banner.cell(row=row_num, column=3, value=values[2])
        cell.fill = REQUIRED_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.font = Font(size=9, bold=True)
        cell.border = THIN_BORDER

        cell = ws_banner.cell(row=row_num, column=4, value=values[3])
        cell.fill = INSTRUCTIONS_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.font = Font(size=9, italic=True)
        cell.border = THIN_BORDER

    # Example data
    examples = [
        ['Total', 'Total Sample', '', ''],
        ['Gender', 'Gender', '', ''],
        ['AgeGroup', 'Age Group', '', ''],
        ['Region', 'Region', '', ''],
    ]

    current_row = 4
    for example in examples:
        add_data_row(ws_banner, example, current_row, EXAMPLE_FILL)
        current_row += 1

    set_column_widths(ws_banner, {
        'A': 20, 'B': 20, 'C': 12, 'D': 50
    })

    # Settings sheet
    ws_settings = wb.create_sheet("Settings")

    headers = ['Setting', 'Value', 'Required?', 'Valid Values', 'Description']
    add_header_row(ws_settings, headers)

    settings_data = [
        ('project_name', 'Customer Satisfaction Tracker', 'Required', 'Any text', 'Project name for output filename'),
        ('decimal_places_ratings', '1', 'Required', '0-3', 'Decimal places for mean ratings and averages'),
        ('show_significance', 'Y', 'Required', 'Y/N or TRUE/FALSE', 'Enable significance testing for wave comparisons'),
        ('alpha', '0.05', 'Required', '0.01, 0.05, 0.10', 'Significance level (0.05 = 95% confidence)'),
        ('minimum_base', '30', 'Required', 'Numeric > 0', 'Minimum sample size for significance testing'),
        ('decimal_separator', '.', 'Required', '. or ,', 'Decimal separator: . (US/UK) or , (European)'),
    ]

    current_row = 2
    for setting, value, required, valid, desc in settings_data:
        ws_settings.cell(row=current_row, column=1, value=setting)
        ws_settings.cell(row=current_row, column=2, value=value)

        req_cell = ws_settings.cell(row=current_row, column=3, value=required)
        req_cell.fill = REQUIRED_FILL if required == 'Required' else OPTIONAL_FILL
        req_cell.font = Font(bold=True, size=9)

        ws_settings.cell(row=current_row, column=4, value=valid)
        ws_settings.cell(row=current_row, column=5, value=desc)

        for col in range(1, 6):
            ws_settings.cell(row=current_row, column=col).border = THIN_BORDER
            ws_settings.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

        current_row += 1

    set_column_widths(ws_settings, {
        'A': 25, 'B': 20, 'C': 12, 'D': 20, 'E': 50
    })

    # Save
    output_path = '/home/user/Turas/templates/Tracker_Config_Template_Annotated.xlsx'
    wb.save(output_path)
    print(f"✓ Created: {output_path}")


def create_tracker_question_mapping_template_annotated():
    """Create annotated Tracker Question Mapping template"""
    print("Creating Tracker_Question_Mapping_Template_Annotated.xlsx...")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Instructions sheet
    create_instructions_sheet(wb, "Tracker Question Mapping Template",
        "This template maps question codes across survey waves when questions move or are renumbered. "
        "It allows TurasTracker to follow the same question even when its code changes.",
        [
            {
                'title': 'How to Use This Template',
                'items': [
                    'Add one row per tracked question in the QuestionMap sheet',
                    'Add one column per wave (Wave1, Wave2, Wave3, etc.)',
                    'Enter the wave-specific question code in each wave column',
                    'Leave blank if question not asked in that wave',
                    'For composites, list source questions in SourceQuestions column'
                ]
            },
            {
                'title': 'Question Type Guide',
                'items': [
                    'Rating: Scale questions (1-5, 1-10, etc.) - reports mean/average',
                    'NPS: Net Promoter Score (0-10 scale) - reports NPS score, % promoters/passives/detractors',
                    'SingleChoice: Single-select categorical - reports % for each option',
                    'Composite: Calculated metric combining multiple questions'
                ]
            },
            {
                'title': 'Common Scenarios',
                'items': [
                    'Question moved: Q10 → Q11 → Q12 (same question, different codes)',
                    'Question added: Leave Wave1 blank, fill Wave2 onwards',
                    'Question removed: Fill Wave1-Wave2, leave Wave3 blank',
                    'Question unchanged: Same code across all waves'
                ]
            }
        ])

    # QuestionMap sheet
    ws = wb.create_sheet("QuestionMap")

    headers = [
        'QuestionCode',
        'QuestionText',
        'QuestionType',
        'Wave1',
        'Wave2',
        'Wave3',
        'Wave4',
        'SourceQuestions',
        'Required?',
        'Valid Types',
        'Description'
    ]
    add_header_row(ws, headers)

    # Documentation rows
    doc_rows = [
        ['', '', '', '', '', '', '', '', 'Required', 'Any unique code', 'Standard tracking code (used in TrackedQuestions)'],
        ['', '', '', '', '', '', '', '', 'Required', '', 'Question wording (for documentation)'],
        ['', '', '', '', '', '', '', '', 'Required', 'Rating/NPS/SingleChoice/Composite', 'Type of question (determines analysis method)'],
        ['', '', '', '', '', '', '', '', 'Optional', 'Question code or blank', 'Question code in Wave 1 data (blank if not asked)'],
        ['', '', '', '', '', '', '', '', 'Optional', 'Question code or blank', 'Question code in Wave 2 data (blank if not asked)'],
        ['', '', '', '', '', '', '', '', 'Optional', 'Question code or blank', 'Question code in Wave 3 data (blank if not asked)'],
        ['', '', '', '', '', '', '', '', 'Optional', 'Question code or blank', 'Question code in Wave 4 data (blank if not asked)'],
        ['', '', '', '', '', '', '', '', 'For Composite only', 'Comma-separated codes', 'Source questions (e.g., Q_SAT,Q_VALUE)'],
    ]

    for row_num, values in enumerate(doc_rows, 2):
        for col_num in range(1, 9):
            cell = ws.cell(row=row_num, column=col_num, value=values[col_num-1])
            cell.fill = INSTRUCTIONS_FILL
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.font = Font(size=9, italic=True)
            cell.border = THIN_BORDER

        # Required? column
        cell = ws.cell(row=row_num, column=9, value=values[8])
        cell.fill = REQUIRED_FILL if 'Required' in values[8] else OPTIONAL_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.font = Font(size=9, bold=True)
        cell.border = THIN_BORDER

        # Valid Types and Description columns
        for col_num in [10, 11]:
            cell = ws.cell(row=row_num, column=col_num, value=values[col_num-1])
            cell.fill = INSTRUCTIONS_FILL
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.font = Font(size=9, italic=True)
            cell.border = THIN_BORDER

    # Example data
    examples = [
        ['Q_SAT', 'Overall satisfaction (1-10)', 'Rating', 'Q10', 'Q11', 'Q12', 'Q15', '', '', '', ''],
        ['Q_NPS', 'Likelihood to recommend (0-10)', 'NPS', 'Q25', 'Q26', 'Q27', 'Q30', '', '', '', ''],
        ['Q_VALUE', 'Value for money (1-10)', 'Rating', 'Q15', 'Q16', 'Q17', 'Q20', '', '', '', ''],
        ['Q_BRAND', 'Brand preference', 'SingleChoice', 'Q5', 'Q5', 'Q5', 'Q5', '', '', '', ''],
        ['Q_NEW', 'New question added in Wave 2', 'Rating', '', 'Q30', 'Q31', 'Q32', '', '', '', ''],
        ['COMP_OVERALL', 'Overall Score (Composite)', 'Composite', 'COMP', 'COMP', 'COMP', 'COMP', 'Q_SAT,Q_VALUE', '', '', ''],
    ]

    current_row = 10
    for example in examples:
        add_data_row(ws, example, current_row, EXAMPLE_FILL)
        current_row += 1

    set_column_widths(ws, {
        'A': 15, 'B': 35, 'C': 15, 'D': 10, 'E': 10, 'F': 10, 'G': 10,
        'H': 20, 'I': 15, 'J': 25, 'K': 40
    })

    # Save
    output_path = '/home/user/Turas/templates/Tracker_Question_Mapping_Template_Annotated.xlsx'
    wb.save(output_path)
    print(f"✓ Created: {output_path}")


if __name__ == "__main__":
    print("="*60)
    print("TURAS Tracker Templates Generator")
    print("="*60)
    print()

    create_tracker_config_template_annotated()
    create_tracker_question_mapping_template_annotated()

    print()
    print("="*60)
    print("Tracker templates created successfully")
    print("="*60)
