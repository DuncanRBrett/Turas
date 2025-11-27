#!/usr/bin/env python3
"""
Create example conjoint configuration Excel file
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# ==============================================================================
# SHEET 1: Settings
# ==============================================================================

ws_settings = wb.create_sheet("Settings")

# Settings data
settings_data = [
    ["Setting", "Value", "Description"],
    ["analysis_type", "choice", "Analysis type: 'choice' or 'rating'"],
    ["estimation_method", "auto", "Estimation method: 'auto', 'mlogit', 'clogit', or 'hb'"],
    ["baseline_handling", "first_level_zero", "How to handle baseline: 'first_level_zero' or 'all_levels_explicit'"],
    ["choice_type", "single", "Choice type: 'single', 'single_with_none', 'best_worst', 'continuous_sum'"],
    ["none_as_baseline", "FALSE", "Treat none option as baseline level (TRUE/FALSE)"],
    ["none_label", "None of these", "Label for none option if applicable"],
    ["data_file", "examples/sample_cbc_data.csv", "Path to data file (relative to config file or absolute)"],
    ["output_file", "examples/output/example_results.xlsx", "Path to output Excel file"],
    ["respondent_id_column", "resp_id", "Column name for respondent ID"],
    ["choice_set_column", "choice_set_id", "Column name for choice set ID"],
    ["alternative_id_column", "alternative_id", "Column name for alternative ID (optional)"],
    ["chosen_column", "chosen", "Column name for chosen indicator (1=chosen, 0=not chosen)"],
    ["confidence_level", "0.95", "Confidence level for intervals (0-1)"],
    ["generate_market_simulator", "FALSE", "Generate interactive market simulator sheet (TRUE/FALSE)"],
    ["include_diagnostics", "TRUE", "Include detailed diagnostics in output (TRUE/FALSE)"],
]

# Write settings data
for row_idx, row_data in enumerate(settings_data, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_settings.cell(row=row_idx, column=col_idx, value=value)

        # Header formatting
        if row_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

# Set column widths
ws_settings.column_dimensions['A'].width = 30
ws_settings.column_dimensions['B'].width = 40
ws_settings.column_dimensions['C'].width = 60

# Freeze first row
ws_settings.freeze_panes = "A2"

# ==============================================================================
# SHEET 2: Attributes
# ==============================================================================

ws_attributes = wb.create_sheet("Attributes")

# Attributes data
attributes_data = [
    ["AttributeName", "AttributeLabel", "NumLevels", "Level1", "Level2", "Level3", "Level4", "Level5", "Level6"],
    ["Brand", "Brand", 4, "Apple", "Samsung", "Google", "OnePlus", None, None],
    ["Price", "Price", 4, "$299", "$399", "$499", "$599", None, None],
    ["Screen_Size", "Screen Size", 3, "5.5 inches", "6.1 inches", "6.7 inches", None, None, None],
    ["Battery_Life", "Battery Life", 3, "12 hours", "18 hours", "24 hours", None, None, None],
    ["Camera_Quality", "Camera Quality", 3, "Basic", "Good", "Excellent", None, None, None],
]

# Write attributes data
for row_idx, row_data in enumerate(attributes_data, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_attributes.cell(row=row_idx, column=col_idx, value=value)

        # Header formatting
        if row_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

# Set column widths
ws_attributes.column_dimensions['A'].width = 20
ws_attributes.column_dimensions['B'].width = 20
ws_attributes.column_dimensions['C'].width = 15
for col in ['D', 'E', 'F', 'G', 'H', 'I']:
    ws_attributes.column_dimensions[col].width = 15

# Freeze first row
ws_attributes.freeze_panes = "A2"

# ==============================================================================
# SHEET 3: Instructions
# ==============================================================================

ws_instructions = wb.create_sheet("Instructions")

instructions = [
    ["TURAS CONJOINT ANALYSIS - EXAMPLE CONFIGURATION"],
    [""],
    ["This is an example configuration file for a smartphone choice-based conjoint study."],
    [""],
    ["STUDY DESIGN:"],
    ["- Choice-based conjoint (CBC)"],
    ["- 5 attributes with 3-4 levels each"],
    ["- Auto estimation method (tries mlogit first, falls back to clogit)"],
    ["- First level of each attribute used as reference (utility = 0)"],
    [""],
    ["ATTRIBUTES:"],
    ["1. Brand: Apple, Samsung, Google, OnePlus"],
    ["2. Price: $299, $399, $499, $599"],
    ["3. Screen Size: 5.5\", 6.1\", 6.7\""],
    ["4. Battery Life: 12h, 18h, 24h"],
    ["5. Camera Quality: Basic, Good, Excellent"],
    [""],
    ["TO USE THIS EXAMPLE:"],
    ["1. Ensure sample_cbc_data.csv exists in the same directory"],
    ["2. Run: source('modules/conjoint/R/00_main.R')"],
    ["3. Run: results <- run_conjoint_analysis('modules/conjoint/examples/example_config.xlsx')"],
    ["4. Check output in: examples/output/example_results.xlsx"],
    [""],
    ["CUSTOMIZATION:"],
    ["- Modify attribute names and levels in the Attributes sheet"],
    ["- Adjust settings in the Settings sheet"],
    ["- Update data_file path to point to your data"],
    [""],
    ["For more information, see the specification documents in modules/conjoint/"],
]

for row_idx, row_data in enumerate(instructions, start=1):
    ws_instructions.cell(row=row_idx, column=1, value=row_data[0])

ws_instructions.column_dimensions['A'].width = 100

# Save workbook
output_file = "/home/user/Turas/modules/conjoint/examples/example_config.xlsx"
wb.save(output_file)

print(f"✓ Example configuration created: {output_file}")
